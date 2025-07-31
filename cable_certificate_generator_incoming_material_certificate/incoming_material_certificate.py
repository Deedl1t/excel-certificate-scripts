
import re
import os
import openpyxl
import shutil
import time
import xlwings as xw
from datetime import datetime, timedelta
#  先
# 程序开始时记录当前时间
start_time = time.time()

# 全局变量字典，用于存储文件夹名称和对应的Excel文件
global_variables = {}

root_directory = r'yourpath' 

target_excel_path = r'yourpath' 

template_path_1 = r'yourpath'

template_path_2 = r'yourpath'

#货物计算器
def num_calculator(sheet):
    # 从A2开始遍历列A并统计有序的数字行数
    start_row = 2  # 开始行号
    current_value = 1  # 从1开始的数字
    row_count = 0
    goods_num = 0  # 初始化 goods_num 为0

    for row in sheet.iter_rows(min_row=start_row, max_col=1):
        cell = row[0]
        if cell.value == current_value:
            row_count += 1
            row_number = start_row + row_count - 1
            current_value += 1
            cell_value_b = sheet.cell(row=row_number, column=2).value
            
            # 检查G列或H列是否有纯数字
            cell_value_g = sheet.cell(row=row_number, column=7).value
            cell_value_h = sheet.cell(row=row_number, column=8).value
            
            if cell_value_g is not None and str(cell_value_g).isdigit():
                cell_value = int(cell_value_g)
            elif cell_value_h is not None and str(cell_value_h).isdigit():
                cell_value = int(cell_value_h)
            else:
                print(f"第{row_number}行的G列和H列都没有纯数字: G列值: {cell_value_g}, H列值: {cell_value_h}")
                continue  # 如果G列和H列都没有纯数字，跳过该行

            if "光缆" in cell_value_b:
                goods_num += cell_value  # 如果是光缆，累加数量到 goods_num

    return goods_num


#时间模块 大于10000延期两天 小于10000延期一天 节假日则一直推到工作日
def add_delivery_date(start_date_str, goods_num):
    # 节假日列表，这里包括2023年和2024年的节假日和调休日期
    holidays_2023 = ["20230101", "20230102", "20230121", "20230122", "20230123", "20230124", 
                     "20230125", "20230126", "20230127", "20230405", "20230429", "20230430", 
                     "20230501", "20230502", "20230503", "20230622", "20230623", "20230624", 
                     "20230929", "20230930", "20231001", "20231002", "20231003", "20231004", 
                     "20231005", "20231006"]
    holidays_2024 = ["20240101", "20240210", "20240211", "20240212", "20240213", "20240214", 
                     "20240215", "20240216", "20240217", "20240404", "20240405", "20240406", 
                     "20240501", "20240502", "20240503", "20240504", "20240505", "20240610", 
                     "20240915", "20240916", "20240917", "20241001", "20241002", "20241003", 
                     "20241004", "20241005", "20241006", "20241007"]
    holidays_2025 = ["20250101", "20250128", "20250129", "20250130", "20250131", "20250201", 
                     "20250202", "20250203", "20250204", "20250404", "20250405", "20250406", 
                     "20250501", "20250502", "20250503", "20250504", "20250505", "20250531", 
                     "20250601", "20250602", "20251001", "20251002", "20251003", "20251004", 
                     "20251005", "20251006", "20251007", "20251008"]
    # 添加周末正常上班日期
    working_weekends_2023 = ["20230128", "20230129", "20230423", "20230506", "20230625", "20231007", "20231008"]
    working_weekends_2024 = ["20240204", "20240218", "20240407", "20240428", "20240511", "20240914", "20240929", "20241012"]
    working_weekends_2025 = ["20250126", "20250208", "20250427", "20250928", "20251011"]
    # 将输入的日期字符串转换为日期对象
    start_date = datetime.strptime(start_date_str, '%Y%m%d')
    # 判断是2023年还是2024年，并设置相应的节假日和工作周末列表
    year = start_date.year
    if year == 2023:
        holidays = holidays_2023
        working_weekends = working_weekends_2023
    elif year == 2024:
        holidays = holidays_2024
        working_weekends = working_weekends_2024
    elif year == 2025:
        holidays = holidays_2025
        working_weekends = working_weekends_2025
    # 定义一个变量来追踪已经处理的货物数量
    processed_goods = 0
    # 如果是大量订单，增加两天开始处理
    if goods_num >= 10000:
        start_date += timedelta(days=2)
    else:  # 如果是小量订单，增加一天开始处理
        start_date += timedelta(days=1)
    while processed_goods < goods_num:
        # 检查是否为节假日或周末（除非是工作周末）
        if start_date.strftime('%Y%m%d') in holidays or (start_date.weekday() == 6 and start_date.strftime('%Y%m%d') not in working_weekends):
            start_date += timedelta(days=1)
            continue
        break
    # 处理完当前批货物后，增加一天
    start_date += timedelta(days=1)
    # 返回处理完所有货物的日期（前一天）
    return (start_date - timedelta(days=1)).strftime('%Y/%m/%d').replace('.0', '.')



def rename_and_replace_sheet(goods_num,sheet_num, destination_file, old_prefix, new_prefix, cell_value_f, cell_value_c, variable_name, cell_value):
    
    # 在函数开始时打开工作簿
    creat_file_open = openpyxl.load_workbook(destination_file)
    # 构造旧工作表名称
    old_sheet_name = f"{old_prefix}{sheet_num}"
    new_sheet_name = f"{new_prefix}{sheet_num}"

    print(f"正在处理工作表：{old_sheet_name}")
    
    start_date_str = variable_name[3:11]
    extention_time = add_delivery_date(start_date_str, goods_num)
    
    # 检查工作表是否存在
    if old_sheet_name in creat_file_open.sheetnames:
        sheet = creat_file_open[old_sheet_name]

        # 遍历工作表的所有单元格并替换文本
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == None:
                    continue
                #print(f"当前单元格值：{cell.value}")  # 打印当前单元格值
                if cell.value == "NAME1":
                    cell.value = variable_name #需求单号&订单编号
                elif cell.value == "NAME2":
                    cell.value = cell_value    #产品数量&出货数量 一致
                elif cell.value == "NAME3":
                    cell.value = cell_value    #产品数量&检验数量 一致
                elif cell.value == "NAME4":
                    cell.value = cell_value_c  #规格型号&规格型号
                elif cell.value == "NAME5":
                    cell.value = extention_time      #送检日期
                elif cell.value == "NAME6":
                    cell.value = cell_value_c  #规格型号&规格型号
                
        # 重命名工作表
        sheet.title = new_sheet_name
        print(f"重命名工作表: {old_sheet_name} 为 {new_sheet_name}")
        # 保存更改
        creat_file_open.save(destination_file)
        creat_file_open.close()
        print("更改已保存。")
    else:
        print(f"工作表 {old_sheet_name} 不存在。")
        

def copy_excel_file(source_path, destination_path):
    try:
        shutil.copyfile(source_path, destination_path)
        print(f"文件已成功复制到：{destination_path}")
    except Exception as e:
        print(f"复制文件时出错：{e}")


def number_line(sheet, folder_name, file_name,destination_file,goods_num):
    # 从A8开始遍历列A并统计有序的数字行数
    start_row = 2  # 开始行号
    column_name = 'A'  # 列名
    current_value = 1  # 从1开始的数字
    row_count = 0
    sheet_num_am = 1   #航空插头公头sheet表格默认开始序号
    sheet_num_af = 1   #航空插头母头sheet表格默认开始序号

    print(f"文件夹名称：{folder_name}, 文件名：{file_name}, Sheet1（合同）")
        
    for row in sheet.iter_rows(min_row=start_row, max_col=1):
        cell = row[0]
        if cell.value == current_value:  # 如果单元格中的值等于当前值
            row_count += 1    
            row_number = start_row + row_count - 1  # 计算当前行号
            current_value += 1  # 增加当前值以匹配下一个数字
            # 根据不同的值执行不同的操作
            cell_value_b = sheet.cell(row=row_number, column=2).value  # B列的值 名称
            cell_value_c = sheet.cell(row=row_number, column=3).value  # C列的值 规格型号
            
            # 判断G列或H列是否有纯数字
            cell_value_g = sheet.cell(row=row_number, column=7).value
            cell_value_h = sheet.cell(row=row_number, column=8).value
            
            if cell_value_g is not None and str(cell_value_g).isdigit():
                cell_value = int(cell_value_g)
                cell_value_f = sheet.cell(row=row_number, column=5).value  # E列的值 品牌
            elif cell_value_h is not None and str(cell_value_h).isdigit():
                cell_value = int(cell_value_h)
                cell_value_f = sheet.cell(row=row_number, column=6).value  # F列的值 品牌
            else:
                print(f"G or H no value! It's {row_number} row")
                continue  # 如果G列和H列都没有纯数字，跳过该行

            if "光缆" in cell_value_b:
                if "多模" in cell_value_c:
                    continue
                elif "单模" in cell_value_c:
                    continue
            elif "机箱" in cell_value_b:
                if "1U" in cell_value_c:
                    continue
                else:
                    print("有其他类型的机箱")
            elif "连接器" in cell_value_b:
                continue
            elif "适配器" in cell_value_b:
                if "单联" in cell_value_c:
                    continue
                elif "双联" in cell_value_c:
                    continue     
                elif "四联" in cell_value_c:
                    continue     
            elif "分支器" in cell_value_b:
                continue
            
            # 判断B列的元素属于哪种产品类型
            if "航空插头" in cell_value_b:
                if "公头" in cell_value_c:
                    old_prefix = "公头"
                    new_prefix = "公头更新"
                    sheet_num = sheet_num_am
                    rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                    print(f"{row_count}:航空插头公头 H{row_number}数据：{cell_value} 公头sheet{sheet_num_am}")
                    sheet_num_am += 1
                    continue
                elif "母头" in cell_value_c:
                    old_prefix = "母头"
                    new_prefix = "母头更新"
                    sheet_num = sheet_num_af
                    rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                    print(f"{row_count}:航空插头母头 H{row_number}数据：{cell_value} 母头sheet{sheet_num_af}")
                    sheet_num_af += 1
                    continue            
            else:
                print(f"error line {row_count}") # 分辨是第几行定位错误地址
        else:
            break  # 如果不是有序的数字，停止遍历
    # 打印有序数字的行数
    print(f"从{column_name}{start_row}开始共有{row_count}行有序数字。\n")


#删除非更新sheet
def delete_specific_sheets_in_directory(directory_path):
    # 遍历目录下的所有文件
    for filename in os.listdir(directory_path):
        # 检查文件是否为 Excel 文件
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            file_path = os.path.join(directory_path, filename)
            print(f"正在处理文件：{file_path}")

            # 加载工作簿
            workbook = openpyxl.load_workbook(file_path)
            # 获取所有工作表名称的列表
            sheet_names = workbook.sheetnames

            # 记录要删除的工作表名称
            sheets_to_delete = []

            # 遍历所有工作表名称
            for sheet_name in sheet_names:
                # 如果工作表名称不是要保留的名称，则记录该工作表名称
                if "更新" not in sheet_name:
                    sheets_to_delete.append(sheet_name)

            # 如果要删除的工作表数量等于现有的工作表数量，则删除整个文件
            if len(sheets_to_delete) >= len(sheet_names):
                workbook.close()
                os.remove(file_path)
                print(f"已删除整个文件：{file_path}")
            else:
                # 删除记录的工作表
                for sheet_name in sheets_to_delete:
                    del workbook[sheet_name]
                    print(f"删除工作表: {sheet_name}")

                # 保存更改
                workbook.save(file_path)
                print(f"已更新工作簿 {file_path}")
                # 关闭工作簿
                workbook.close()

#获得系统当前时间
def get_current_datetime():
    return datetime.now().strftime("%Y%m%d%H%M%S")

# 遍历根目录下的所有文件夹
for folder_name in sorted(os.listdir(root_directory)):
    folder_path = os.path.join(root_directory, folder_name)
    
    # 检查是否是文件夹
    if os.path.isdir(folder_path):
        # 使用正则表达式提取文件夹名称中符合 <FOLDER_CODE_PATTERN> 格式的子字符串
        match = re.search(r"<FOLDER_CODE_PATTERN>", folder_name)
        if match:
            variable_name = match.group()  # 提取匹配的子字符串
            
            # 初始化计数器
            count = 1
            # 统计含有 "物料编码" 文件的数量
            file_count = sum(1 for file_name in os.listdir(folder_path) if "物料编码" in file_name and file_name.endswith('.xlsx'))
            
            # 遍历文件夹下的文件
            for file_name in os.listdir(folder_path):
                if "物料编码" in file_name and file_name.endswith('.xlsx'):
                    # 确定文件前缀
                    prefix = ""
                    #original_excel_path = None  # 初始化为None
                    if "x" in file_name:
                        prefix = "x_"
                        original_excel_path = template_path_1
                    elif "y" in file_name:
                        prefix = "y_"
                        original_excel_path = template_path_2
                        
                    #系统当前时间
                    current_date = get_current_datetime()
                    
                    # 判断是否需要添加计数后缀
                    if file_count > 1:
                        destination_file = os.path.join(target_excel_path, f"{current_date}_{prefix}航空插头合格证{variable_name}-{count}.xlsx")
                    else:
                        destination_file = os.path.join(target_excel_path, f"{current_date}_{prefix}航空插头合格证{variable_name}.xlsx")
                        
                    # 调用函数复制文件
                    copy_excel_file(original_excel_path, destination_file)
                    
                    # 打开Excel文件
                    excel_path = os.path.join(folder_path, file_name)
                    wb = openpyxl.load_workbook(excel_path)
                    sheet = wb.worksheets[0]
                    
                    # 数量计算，天数要用 10000 2天
                    goods_num = num_calculator(sheet)
                    
                    # 打印Sheet1（合同）的H8开始直到结束的数据
                    number_line(sheet, folder_name, file_name, destination_file, goods_num)
                    
                    # 存储新创建的文件路径
                    if file_count > 1:
                        global_variables[f"{variable_name}-{count}"] = destination_file
                    else:
                        global_variables[f"{variable_name}"] = destination_file
                    count += 1

        print(f"所有更改已保存到 {destination_file}")

delete_specific_sheets_in_directory(target_excel_path)

# 遍历文件夹中的所有文件
for filename_trans_pdf in os.listdir(target_excel_path):
    # 检查文件是否为Excel文件
    if filename_trans_pdf.endswith('.xlsx'):
        # 构造完整文件路径
        file_path = os.path.join(target_excel_path, filename_trans_pdf)
        # 构造输出PDF文件路径
        output_pdf_path = os.path.join(target_excel_path, f'{os.path.splitext(filename_trans_pdf)[0]}.pdf')
        
        # 启动 Excel 应用程序
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        
        # 保存为 PDF
        wb.api.ExportAsFixedFormat(0, output_pdf_path)
        
        # 关闭工作簿
        wb.close()
        # 关闭 Excel 应用程序
        app.quit()
        
        print(f'PDF saved to {output_pdf_path}')

# 打印全局变量信息
for variable_name, excel_path in global_variables.items():
    print(f"全局变量名称-文件夹名称：{variable_name}")

# 程序运行结束时记录时间
end_time = time.time()

# 计算总运行时间
total_time = end_time - start_time
total_min = total_time // 60
total_sec = total_time % 60
print(f"程序总运行时间：{total_time}秒")
print(f"程序总运行时间：{total_min}分{total_sec}秒")