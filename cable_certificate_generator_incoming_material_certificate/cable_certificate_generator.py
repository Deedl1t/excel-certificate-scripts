
import os
import re
import openpyxl
import shutil
import xlwings as xw
import time
from datetime import datetime, timedelta

# 程序开始时记录当前时间
start_time = time.time()

# 全局变量字典，用于存储文件夹名称和对应的Excel文件
global_variables = {}
global_plate_num = 1  # 全局变量，用于跟踪盘号

# 根目录文件夹路径 
root_directory = r'yourpath' 

# 目标文件夹路径 
target_excel_path = r'yourpath' 

template_path_1  = r'yourpath'

template_path_2 = r'yourpath'

# 每行光缆数据分割，添加到数组中
def OC_quantity_division(sheet):
    start_row = 2  # 开始行号
    current_value = 1  # 从1开始的数字
    row_count = 0
    OC_qd_Arrays_3d = []
    row_material_oc_num = 0

    # 光缆物料累加
    for row in sheet.iter_rows(min_row=start_row, max_col=1):
        cell = row[0]
        if cell.value == current_value:
            row_count += 1
            row_number = start_row + row_count - 1
            current_value += 1
            cell_value_b = sheet.cell(row=row_number, column=2).value
            cell_value_c = sheet.cell(row=row_number, column=3).value
            cur_row_oc_qd_quantity = 0

            # 检查G列或H列是否有纯数字
            cell_value_g = sheet.cell(row=row_number, column=7).value
            cell_value_h = sheet.cell(row=row_number, column=8).value

            if cell_value_g is not None and str(cell_value_g).isdigit():
                cell_value = int(cell_value_g)
            elif cell_value_h is not None and str(cell_value_h).isdigit():
                cell_value = int(cell_value_h)
            else:
                print(f"第{row_number}行的G列和H列都没有纯数字: G列值: {cell_value_g}, H列值: {cell_value_h}")
                continue  # 如果G列和H列都没有纯数字，跳过该行

            if "光缆" in cell_value_b:
                row_material_oc_num += 1
                OC_qd_Arrays_2d = []  # 初始化二维数组
                if 0 < cell_value <= 5000:
                    OC_qd_Arrays_2d.append(cell_value)
                    cur_row_oc_qd_quantity += 1
                elif cell_value > 5000:
                    while cell_value > 5000:
                        OC_qd_Arrays_2d.append(5000)
                        cell_value -= 5000
                        cur_row_oc_qd_quantity += 1
                    if cell_value > 0:
                        OC_qd_Arrays_2d.append(cell_value)
                        cur_row_oc_qd_quantity += 1
                OC_qd_Arrays_2d.append(row_material_oc_num)            #光缆物料sheet块           -5
                OC_qd_Arrays_2d.append(cur_row_oc_qd_quantity)         #光缆物料分割总数          -4

                # 识别芯数
                core_count = None
                core_counts = ["24芯", "4芯", "6芯", "8芯", "12芯"]
                for count in core_counts:
                    if count in cell_value_c:
                        core_count = count
                        break
                if core_count:
                    OC_qd_Arrays_2d.append(core_count)                  # 光缆芯数                 -3
                else:
                    OC_qd_Arrays_2d.append("未知芯数")  # 如果没有匹配到芯数

                # 识别光缆类型
                if "单模" in cell_value_c:
                    OC_qd_Arrays_2d.append("单模")                      #-1  光缆类型：单模或多模    -2
                elif "多模" in cell_value_c:
                    OC_qd_Arrays_2d.append("多模")
                else:
                    OC_qd_Arrays_2d.append("未知类型")
                OC_qd_Arrays_2d.append(cell_value_c)                   #光缆型号规格              -1                

                OC_qd_Arrays_3d.append(OC_qd_Arrays_2d)  # 将二维数组添加到三维数组中

    return OC_qd_Arrays_3d
    #返回三维数组，有物料光缆的总行数


# 数据填充函数
def data_filling(sheet, data, variable_name, j, goods_num):
    global global_plate_num  # 声明使用全局变量

    specific_value_optical_Specification = data[-1]  # 光缆规格
    specific_value_optical_type = data[-2]  # 光缆类型
    specific_value_optical_core_number = data[-3]  # 光缆芯数
    
    start_date_str = variable_name[3:11]
    extention_time, extention_date_chinese, final_date = add_delivery_date(start_date_str, goods_num)

    # 提取月份和天数
    month_day = f"{final_date.month}{final_date.day:02d}"
    # 转换为整数并乘以 1000
    date_plate_num = int(month_day) * 1000
    
    for row in sheet.range('A1:R21'):  # 假设A1:Z100是需要遍历的范围，可以调整范围
        for cell in row:
            if cell.value is None:
                continue
            if cell.value == "NAME1":
                cell.value = specific_value_optical_Specification  #型号规格
            elif cell.value == "NAME2":
                cell.value = data[j - 1]  # 注意：因为 j 从 1 开始，光缆长度
            elif cell.value == "NAME3":
                cell.value = date_plate_num + global_plate_num  # 盘号
            elif cell.value == "NAME4":
                cell.value = variable_name
            elif cell.value == "NAME5":
                cell.value = extention_date_chinese  # 制造日期
            elif cell.value == "NAME6":
                cell.value = extention_date_chinese  # 测试日期

            # 初始化 ranges_to_fill_none
            ranges_to_fill_none = []

            # 填充单模/多模报告特有的数据
            if specific_value_optical_type == "单模":
                if specific_value_optical_core_number == "24芯":
                    # 处理24芯的情况
                    if cell.value == "NAME7":
                        cell.value = "1310nm"
                    elif cell.value == "NAME8":
                        cell.value = "1550nm"
                    elif cell.value == "NAME9":
                        cell.value = "1310nm"
                    elif cell.value == "NAME10":
                        cell.value = "1550nm"
                    # 填充指定范围内的单元格
                    ranges_to_fill_318 = ["NAME11", "NAME12", "NAME13", "NAME14", "NAME19", "NAME20", "NAME23", "NAME24", 
                                          "NAME27", "NAME28", "NAME29", "NAME30", "NAME35", "NAME36", "NAME37", "NAME38", 
                                          "NAME39", "NAME40", "NAME41", "NAME42", "NAME43", "NAME44", "NAME45", "NAME46"]
                    ranges_to_fill_180 = ["NAME15", "NAME16", "NAME17", "NAME18", "NAME21", "NAME22", "NAME25", "NAME26", 
                                          "NAME31", "NAME32", "NAME33", "NAME34", "NAME47", "NAME48", "NAME49", "NAME50", 
                                          "NAME51", "NAME52", "NAME53", "NAME54", "NAME55", "NAME56", "NAME57", "NAME58"]
                elif specific_value_optical_core_number == "4芯":
                    if cell.value == "NAME7":
                        cell.value = "1310nm"
                    elif cell.value == "NAME8":
                        cell.value = "1550nm"
                    elif cell.value == "NAME9":
                        cell.value = None
                    elif cell.value == "NAME10":
                        cell.value = None
                    ranges_to_fill_318 = ["NAME11", "NAME12", "NAME13", "NAME14"]
                    ranges_to_fill_180 = ["NAME15", "NAME16", "NAME17", "NAME18"]
                    ranges_to_fill_none = [f"NAME{i}" for i in range(19, 59)]
                elif specific_value_optical_core_number == "6芯":
                    if cell.value == "NAME7":
                        cell.value = "1310nm"
                    elif cell.value == "NAME8":
                        cell.value = "1550nm"
                    elif cell.value == "NAME9":
                        cell.value = None
                    elif cell.value == "NAME10":
                        cell.value = None
                    ranges_to_fill_318 = ["NAME11", "NAME12", "NAME13", "NAME14", "NAME19", "NAME20"]
                    ranges_to_fill_180 = ["NAME15", "NAME16", "NAME17", "NAME18", "NAME21", "NAME22"]
                    ranges_to_fill_none = [f"NAME{i}" for i in range(23, 59)]
                elif specific_value_optical_core_number == "8芯":
                    if cell.value == "NAME7":
                        cell.value = "1310nm"
                    elif cell.value == "NAME8":
                        cell.value = "1550nm"
                    elif cell.value == "NAME9":
                        cell.value = None
                    elif cell.value == "NAME10":
                        cell.value = None
                    ranges_to_fill_318 = ["NAME11", "NAME12", "NAME13", "NAME14", "NAME19", "NAME20", "NAME23", "NAME24"]
                    ranges_to_fill_180 = ["NAME15", "NAME16", "NAME17", "NAME18", "NAME21", "NAME22", "NAME25", "NAME26"]
                    ranges_to_fill_none = [f"NAME{i}" for i in range(27, 59)]
                elif specific_value_optical_core_number == "12芯":
                    if cell.value == "NAME7":
                        cell.value = "1310nm"
                    elif cell.value == "NAME8":
                        cell.value = "1550nm"
                    elif cell.value == "NAME9":
                        cell.value = None
                    elif cell.value == "NAME10":
                        cell.value = None
                    ranges_to_fill_318 = ["NAME11", "NAME12", "NAME13", "NAME14", "NAME19", "NAME20", "NAME23", "NAME24", "NAME27", "NAME28", "NAME29", "NAME30"]
                    ranges_to_fill_180 = ["NAME15", "NAME16", "NAME17", "NAME18", "NAME21", "NAME22", "NAME25", "NAME26", "NAME31", "NAME32", "NAME33", "NAME34"]
                    ranges_to_fill_none = [f"NAME{i}" for i in range(35, 59)]
                
                if cell.value in ranges_to_fill_318:
                    cell.value = "=0.318+0.009*RAND()"
                elif cell.value in ranges_to_fill_180:
                    cell.value = "=0.18+0.009*RAND()"
                elif cell.value in ranges_to_fill_none:
                    cell.value = None

            elif specific_value_optical_type == "多模":
                if specific_value_optical_core_number == "24芯":
                    if cell.value == "NAME7":
                        cell.value = "850nm"
                    elif cell.value == "NAME8":
                        cell.value = "1300nm"
                    elif cell.value == "NAME9":
                        cell.value = "850nm"
                    elif cell.value == "NAME10":
                        cell.value = "1300nm"
                    ranges_to_fill_318 = ["NAME11", "NAME12", "NAME13", "NAME14", "NAME19", "NAME20", "NAME23", "NAME24", 
                                          "NAME27", "NAME28", "NAME29", "NAME30", "NAME35", "NAME36", "NAME37", "NAME38", 
                                          "NAME39", "NAME40", "NAME41", "NAME42", "NAME43", "NAME44", "NAME45", "NAME46"]
                    ranges_to_fill_180 = ["NAME15", "NAME16", "NAME17", "NAME18", "NAME21", "NAME22", "NAME25", "NAME26", 
                                          "NAME31", "NAME32", "NAME33", "NAME34", "NAME47", "NAME48", "NAME49", "NAME50", 
                                          "NAME51", "NAME52", "NAME53", "NAME54", "NAME55", "NAME56", "NAME57", "NAME58"]
                elif specific_value_optical_core_number == "4芯":
                    if cell.value == "NAME7":
                        cell.value = "850nm"
                    elif cell.value == "NAME8":
                        cell.value = "1300nm"
                    elif cell.value == "NAME9":
                        cell.value = None
                    elif cell.value == "NAME10":
                        cell.value = None
                    ranges_to_fill_318 = ["NAME11", "NAME12", "NAME13", "NAME14"]
                    ranges_to_fill_180 = ["NAME15", "NAME16", "NAME17", "NAME18"]
                    ranges_to_fill_none = [f"NAME{i}" for i in range(19, 59)]
                elif specific_value_optical_core_number == "6芯":
                    if cell.value == "NAME7":
                        cell.value = "850nm"
                    elif cell.value == "NAME8":
                        cell.value = "1300nm"
                    elif cell.value == "NAME9":
                        cell.value = None
                    elif cell.value == "NAME10":
                        cell.value = None
                    ranges_to_fill_318 = ["NAME11", "NAME12", "NAME13", "NAME14", "NAME19", "NAME20"]
                    ranges_to_fill_180 = ["NAME15", "NAME16", "NAME17", "NAME18", "NAME21", "NAME22"]
                    ranges_to_fill_none = [f"NAME{i}" for i in range(23, 59)]
                elif specific_value_optical_core_number == "8芯":
                    if cell.value == "NAME7":
                        cell.value = "850nm"
                    elif cell.value == "NAME8":
                        cell.value = "1300nm"
                    elif cell.value == "NAME9":
                        cell.value = None
                    elif cell.value == "NAME10":
                        cell.value = None
                    ranges_to_fill_318 = ["NAME11", "NAME12", "NAME13", "NAME14", "NAME19", "NAME20", "NAME23", "NAME24"]
                    ranges_to_fill_180 = ["NAME15", "NAME16", "NAME17", "NAME18", "NAME21", "NAME22", "NAME25", "NAME26"]
                    ranges_to_fill_none = [f"NAME{i}" for i in range(27, 59)]
                elif specific_value_optical_core_number == "12芯":
                    if cell.value == "NAME7":
                        cell.value = "850nm"
                    elif cell.value == "NAME8":
                        cell.value = "1300nm"
                    elif cell.value == "NAME9":
                        cell.value = None
                    elif cell.value == "NAME10":
                        cell.value = None
                    ranges_to_fill_318 = ["NAME11", "NAME12", "NAME13", "NAME14", "NAME19", "NAME20", "NAME23", "NAME24", "NAME27", "NAME28", "NAME29", "NAME30"]
                    ranges_to_fill_180 = ["NAME15", "NAME16", "NAME17", "NAME18", "NAME21", "NAME22", "NAME25", "NAME26", "NAME31", "NAME32", "NAME33", "NAME34"]
                    ranges_to_fill_none = [f"NAME{i}" for i in range(35, 59)]
                
                if cell.value in ranges_to_fill_318:
                    cell.value = "=3+0.02*RAND()"
                elif cell.value in ranges_to_fill_180:
                    cell.value = "=1.1+0.02*RAND()"
                elif cell.value in ranges_to_fill_none:
                    cell.value = None



# 创建并填充光缆合格证的sheet
def Creat_sheet_for_OCQD3(destination_file, OC_qd_Arrays_3d, goods_num):
    global global_plate_num  # 声明使用全局变量
    
    # 打开Excel文件
    app = xw.App(visible=False)
    wb = app.books.open(destination_file)
    
    # 遍历OC_qd_Arrays_3d
    for i, data in enumerate(OC_qd_Arrays_3d):
        specific_value_optical_Specification = data[-1]
        specific_value_optical_type = data[-2]
        specific_value_optical_core_number = data[-3]
        specific_value_optical_copy_sheet_number = data[-4]
        specific_value_optical_material_row_number = data[-5]
        
        # 确定要复制的sheet表
        if specific_value_optical_type == "单模":
            base_sheet_name = "单模"
            report_sheet_name = "单模报告"
        elif specific_value_optical_type == "多模":
            base_sheet_name = "多模"
            report_sheet_name = "多模报告"
        else:
            print(f"未知的光缆类型: {specific_value_optical_type}")
            continue
        
        # 复制sheet表并重命名
        for j in range(1, specific_value_optical_copy_sheet_number + 1):
            new_sheet_name = f"更新_{specific_value_optical_core_number}{specific_value_optical_type}{specific_value_optical_material_row_number}-{j}"
            new_report_sheet_name = f"更新_{specific_value_optical_core_number}{specific_value_optical_type}报告{specific_value_optical_material_row_number}-{j}"
            
            # 复制并重命名基本sheet表
            base_sheet = wb.sheets[base_sheet_name]
            new_base_sheet = base_sheet.copy(after=wb.sheets[-1])
            new_base_sheet.name = new_sheet_name
            data_filling(new_base_sheet, data, variable_name, j, goods_num)
            
            # 复制并重命名报告sheet表
            report_sheet = wb.sheets[report_sheet_name]
            new_report_sheet = report_sheet.copy(after=wb.sheets[-1])
            new_report_sheet.name = new_report_sheet_name
            data_filling(new_report_sheet, data, variable_name, j, goods_num)
            
            print(f"创建新的sheet表: {new_sheet_name} 和 {new_report_sheet_name}")
            
            global_plate_num += 1  # 增加盘号

    # 保存并关闭工作簿
    wb.save()
    wb.close()
    app.quit()


# 货物计算器
def num_calculator(sheet):
    # 从A2开始遍历列A并统计有序的数字行数
    start_row = 2  # 开始行号
    current_value = 1  # 从1开始的数字
    row_count = 0
    goods_num = 0  # 初始化 goods_num 为0

    for row in sheet.iter_rows(min_row=start_row, max_col=1):
        cell = row[0]
        if cell.value == current_value:
            row_count += 1
            row_number = start_row + row_count - 1
            current_value += 1
            cell_value_b = sheet.cell(row=row_number, column=2).value
            
            # 检查G列或H列是否有纯数字
            cell_value_g = sheet.cell(row=row_number, column=7).value
            cell_value_h = sheet.cell(row=row_number, column=8).value
            
            if cell_value_g is not None and str(cell_value_g).isdigit():
                cell_value = int(cell_value_g)
            elif cell_value_h is not None and str(cell_value_h).isdigit():
                cell_value = int(cell_value_h)
            else:
                print(f"第{row_number}行的G列和H列都没有纯数字: G列值: {cell_value_g}, H列值: {cell_value_h}")
                continue  # 如果G列和H列都没有纯数字，跳过该行

            if "光缆" in cell_value_b:
                goods_num += cell_value  # 如果是光缆，累加数量到 goods_num

    return goods_num


# 时间模块 大于10000延期两天 小于10000延期一天 节假日则一直推到工作日
def add_delivery_date(start_date_str, goods_num):
    # 节假日列表，这里包括2023年和2024年的节假日和调休日期
    holidays_2023 = ["20230101", "20230102", "20230121", "20230122", "20230123", "20230124", 
                     "20230125", "20230126", "20230127", "20230405", "20230429", "20230430", 
                     "20230501", "20230502", "20230503", "20230622", "20230623", "20230624", 
                     "20230929", "20230930", "20231001", "20231002", "20231003", "20231004", 
                     "20231005", "20231006"]
    holidays_2024 = ["20240101", "20240210", "20240211", "20240212", "20240213", "20240214", 
                     "20240215", "20240216", "20240217", "20240404", "20240405", "20240406", 
                     "20240501", "20240502", "20240503", "20240504", "20240505", "20240610", 
                     "20240915", "20240916", "20240917", "20241001", "20241002", "20241003", 
                     "20241004", "20241005", "20241006", "20241007"]
    holidays_2025 = ["20250101", "20250128", "20250129", "20250130", "20250131", "20250201", 
                     "20250202", "20250203", "20250204", "20250404", "20250405", "20250406", 
                     "20250501", "20250502", "20250503", "20250504", "20250505", "20250531", 
                     "20250601", "20250602", "20251001", "20251002", "20251003", "20251004", 
                     "20251005", "20251006", "20251007", "20251008"]
    # 添加周末正常上班日期
    working_weekends_2023 = ["20230128", "20230129", "20230423", "20230506", "20230625", "20231007", "20231008"]
    working_weekends_2024 = ["20240204", "20240218", "20240407", "20240428", "20240511", "20240914", "20240929", "20241012"]
    working_weekends_2025 = ["20250126", "20250208", "20250427", "20250928", "20251011"]
    # 将输入的日期字符串转换为日期对象
    start_date = datetime.strptime(start_date_str, '%Y%m%d')
    # 判断是2023年还是2024年，并设置相应的节假日和工作周末列表
    year = start_date.year
    if year == 2023:
        holidays = holidays_2023
        working_weekends = working_weekends_2023
    elif year == 2024:
        holidays = holidays_2024
        working_weekends = working_weekends_2024
    elif year == 2025:
        holidays = holidays_2025
        working_weekends = working_weekends_2025
    # 定义一个变量来追踪已经处理的货物数量
    processed_goods = 0
    # 如果是大量订单，增加两天开始处理
    if goods_num >= 10000:
        start_date += timedelta(days=2)
    else:  # 如果是小量订单，增加一天开始处理
        start_date += timedelta(days=1)
    while processed_goods < goods_num:
        # 检查是否为节假日或周末（除非是工作周末）
        if start_date.strftime('%Y%m%d') in holidays or (start_date.weekday() == 6 and start_date.strftime('%Y%m%d') not in working_weekends):
            start_date += timedelta(days=1)
            continue
        break
    # 处理完当前批货物后，增加一天
    start_date += timedelta(days=1)
    # 返回处理完所有货物的日期（前一天）
    final_date = start_date - timedelta(days=1)
    extention_time = final_date.strftime('%Y.%m.%d').replace('.0', '.')
    
    # 自定义去掉前导零的日期格式
    year = final_date.year
    month = final_date.month
    day = final_date.day
    extention_date_chinese = f"{year}年{month}月{day}日"
    
    return extention_time, extention_date_chinese, final_date  # 返回final_date

def copy_excel_file(source_path, destination_path):
    try:
        shutil.copyfile(source_path, destination_path)
        print(f"文件已成功复制到：{destination_path}")
    except Exception as e:
        print(f"复制文件时出错：{e}")

# 删除非更新sheet
def delete_specific_sheets_in_directory(directory_path):
    # 遍历目录下的所有文件
    for filename in os.listdir(directory_path):
        # 检查文件是否为 Excel 文件
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            file_path = os.path.join(directory_path, filename)
            print(f"正在处理文件：{file_path}")

            # 加载工作簿
            workbook = openpyxl.load_workbook(file_path)
            # 获取所有工作表名称的列表
            sheet_names = workbook.sheetnames

            # 记录要删除的工作表名称
            sheets_to_delete = []

            # 遍历所有工作表名称
            for sheet_name in sheet_names:
                # 如果工作表名称不是要保留的名称，则记录该工作表名称
                if "更新" not in sheet_name:
                    sheets_to_delete.append(sheet_name)

            # 如果要删除的工作表数量等于现有的工作表数量，则删除整个文件
            if len(sheets_to_delete) >= len(sheet_names):
                workbook.close()
                os.remove(file_path)
                print(f"已删除整个文件：{file_path}")
            else:
                # 删除记录的工作表
                for sheet_name in sheets_to_delete:
                    del workbook[sheet_name]
                    print(f"删除工作表: {sheet_name}")

                # 保存更改
                workbook.save(file_path)
                print(f"已更新工作簿 {file_path}")
                # 关闭工作簿
                workbook.close()

#获得系统当前时间
def get_current_datetime():
    return datetime.now().strftime("%Y%m%d%H%M%S")

# 遍历根目录下的所有文件夹
for folder_name in sorted(os.listdir(root_directory)):
    folder_path = os.path.join(root_directory, folder_name)
    
    # 检查是否是文件夹
    if os.path.isdir(folder_path):
        # 使用正则表达式提取文件夹名称中符合 <CODE_PATTERN> 格式的子字符串
        match = re.search(r"<CODE_PATTERN>", folder_name)
        if match:
            variable_name = match.group()  # 提取匹配的子字符串
            
            # 初始化计数器
            count = 1
            # 统计含有 "物料编码" 文件的数量
            file_count = sum(1 for file_name in os.listdir(folder_path) if "物料编码" in file_name and file_name.endswith('.xlsx'))
            
            # 遍历文件夹下的文件
            for file_name in os.listdir(folder_path):
                if "物料编码" in file_name and file_name.endswith('.xlsx'):
                    # 确定文件前缀
                    prefix = ""
                    if "x" in file_name:
                        prefix = "x_"
                        original_excel_path = template_path_1
                    elif "y" in file_name:
                        prefix = "y_"
                        original_excel_path = template_path_2
                    
                    #系统当前时间
                    current_date = get_current_datetime()
                    
                    # 判断是否需要添加计数后缀
                    if file_count > 1:
                        destination_file = os.path.join(target_excel_path, f"{current_date}_{prefix}光缆合格证{variable_name}-{count}.xlsx")
                    else:
                        destination_file = os.path.join(target_excel_path, f"{current_date}_{prefix}光缆合格证{variable_name}.xlsx")
                        
                    # 调用函数复制文件
                    copy_excel_file(original_excel_path, destination_file)
                    
                    # 打开Excel文件
                    excel_path = os.path.join(folder_path, file_name)
                    wb = openpyxl.load_workbook(excel_path)
                    sheet = wb.worksheets[0]
                    
                    # 数量计算，天数要用 10000 2天
                    goods_num = num_calculator(sheet)
                    
                    # 每行光缆数据分割，添加到数组中
                    OC_qd_Arrays_3d = OC_quantity_division(sheet)
                    
                    # 通过OC_quantity_division(sheet)返回值来进行excel的sheet表复制，并替换sheet表的表名
                    Creat_sheet_for_OCQD3(destination_file, OC_qd_Arrays_3d, goods_num)
                    
                    # 存储新创建的文件路径
                    if file_count > 1:
                        global_variables[f"{variable_name}-{count}"] = destination_file
                    else:
                        global_variables[f"{variable_name}"] = destination_file
                    count += 1

            # 重置全局盘号计数器
            global_plate_num = 1

        print(f"所有更改已保存到 {destination_file}")

delete_specific_sheets_in_directory(target_excel_path)

# 遍历文件夹中的所有文件
for filename_trans_pdf in os.listdir(target_excel_path):
    # 检查文件是否为Excel文件
    if filename_trans_pdf.endswith('.xlsx'):
        # 构造完整文件路径
        file_path = os.path.join(target_excel_path, filename_trans_pdf)
        # 构造输出PDF文件路径
        output_pdf_path = os.path.join(target_excel_path, f'{os.path.splitext(filename_trans_pdf)[0]}.pdf')
        
        # 启动 Excel 应用程序
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        
        # 保存为 PDF
        wb.api.ExportAsFixedFormat(0, output_pdf_path)
        
        # 关闭工作簿
        wb.close()
        # 关闭 Excel 应用程序
        app.quit()
        
        print(f'PDF saved to {output_pdf_path}')

# 打印全局变量信息
for variable_name, excel_path in global_variables.items():
    print(f"全局变量名称-文件夹名称：{variable_name}")

# 程序运行结束时记录时间
end_time = time.time()

# 计算总运行时间
total_time = end_time - start_time
total_min = total_time // 60
total_sec = total_time % 60
print(f"程序总运行时间：{total_time}秒")
print(f"程序总运行时间：{total_min}分{total_sec}秒")
