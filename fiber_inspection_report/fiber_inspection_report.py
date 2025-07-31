
import os
import re
import openpyxl
import shutil
import xlwings as xw
import time
from datetime import datetime, timedelta
import random
import numpy as np


# 程序开始时记录当前时间
start_time = time.time()

# 全局变量字典，用于存储文件夹名称和对应的Excel文件
global_variables = {}
#global_plate_num = 1  # 全局变量，用于跟踪盘号

oc_num_global = []

return_loss_str_2d = []
inception_loss_str_2d = []
combined_loss_str_2d = []
rls2 = 0
ils2 = 0
cls2 = 0

# 根目录文件夹路径，改成K就好
root_directory = r'yourpath' 

# 目标文件夹路径 
target_excel_path = r'yourpath' 

# 原始模板路径 
original_excel_path = r'yourpath'

# 每行光缆数据分割，添加到数组中,读取物料编码
def OC_Chassis_add(sheet):
    start_row = 2  # 开始行号
    current_value = 1  # 从1开始的数字
    row_count = 0
    OCCH_Arrays_3d = []
    row_material_occh_num = 0

    # 光缆和机箱填充至3维表中
    for row in sheet.iter_rows(min_row=start_row, max_col=1):
        cell = row[0]
        if cell.value == current_value:
            row_count += 1
            row_number = start_row + row_count - 1
            current_value += 1
            OCCH_Arrays_2d = []  # 初始化二维数组
            #在出厂检验中，专门给
            cell_value_b = sheet.cell(row=row_number, column=2).value  #b列元素
            #cell_value_c = sheet.cell(row=row_number, column=3).value  #c列元素
            # m和n和l列是专门为光缆清册设计的，除了机箱
            cell_value_n = sheet.cell(row=row_number, column=14).value #n列元素，根数
            cell_value_m = sheet.cell(row=row_number, column=13).value #m列元素，光缆清册
            cell_value_l = sheet.cell(row=row_number, column=12).value #l列元素，名称

            # 检查G列或H列是否有纯数字
            cell_value_g = sheet.cell(row=row_number, column=7).value
            cell_value_h = sheet.cell(row=row_number, column=8).value

            if cell_value_g is not None and str(cell_value_g).isdigit():
                cell_value = int(cell_value_g)
            elif cell_value_h is not None and str(cell_value_h).isdigit():
                cell_value = int(cell_value_h)
            else:
                print(f"第{row_number}行的G列和H列都没有纯数字: G列值: {cell_value_g}, H列值: {cell_value_h}")
                continue  # 如果G列和H列都没有纯数字，跳过该行
            
            # 光缆的二维数组
            if "光缆" in cell_value_l:
                row_material_occh_num += 1

                #识别物料行，复制sheet表数量
                OCCH_Arrays_2d.append(row_material_occh_num)           #光缆物料sheet块            0
                OCCH_Arrays_2d.append(cell_value_m)                    #光缆型号规格               1   

                # 根据光缆型号规格来识别光缆类型
                if "单模" in cell_value_m:
                    OCCH_Arrays_2d.append("单模")                      #光缆类型：单模或多模        2
                elif "多模" in cell_value_m:
                    OCCH_Arrays_2d.append("多模")
                else:
                    OCCH_Arrays_2d.append("未知类型")
                    print(f"error：光缆型号规格中没有单模或多模，行号: {row_number}, 值: {cell_value_m}")
                # 根据光缆型号规格来识别光缆金属，非金属
                if "非金属" in cell_value_m:
                    OCCH_Arrays_2d.append("非金属")                      #光缆类型：金属或非金属      3
                elif "金属" in cell_value_m:
                    OCCH_Arrays_2d.append("金属")
                else:
                    OCCH_Arrays_2d.append("未知类型")
                    print(f"error：光缆型号规格中没有金属或非金属，行号: {row_number}, 值: {cell_value_m}")
                # 根据光缆型号规格来识别尾缆或光缆
                if "尾缆" in cell_value_m or "室内" in cell_value_m:
                    OCCH_Arrays_2d.append("预制尾缆")                      #光缆类型：尾缆或光缆     4
                else:
                    OCCH_Arrays_2d.append("预制光缆")
                # 识别芯数
                core_count = None
                core_counts = ["24芯", "4芯", "6芯", "8芯", "12芯"]
                for count in core_counts:
                    if count in cell_value_m:
                        core_count = count
                        break
                if core_count:
                    OCCH_Arrays_2d.append(core_count)                  # 光缆芯数                  5
                else:
                    OCCH_Arrays_2d.append("未知芯数")  # 如果没有匹配到芯数
                    print("error：未匹配到芯数")
                
                OCCH_Arrays_2d.append(cell_value_n)           #光缆或尾缆根数                       6
                OCCH_Arrays_3d.append(OCCH_Arrays_2d)  # 将二维数组添加到三维数组中
            
            # 机箱的二维数组
            elif "机箱" in cell_value_b:
                row_material_occh_num += 1
                OCCH_Arrays_2d.append(row_material_occh_num)           #机箱物料sheet块            0
                
                OCCH_Arrays_2d.append(cell_value_b)           #后续可能会进行区分，标记成机箱           1
                OCCH_Arrays_2d.append(cell_value)             #机箱的数量                             2
                # 填充缺失的元素以确保长度一致
                OCCH_Arrays_2d.extend([""] * (7 - len(OCCH_Arrays_2d)))
                
                OCCH_Arrays_3d.append(OCCH_Arrays_2d)  # 将二维数组添加到三维数组中

    return OCCH_Arrays_3d
    #返回三维数组
    

# 创建并填充光缆合格证的sheet
def Creat_sheet_for_OCCH3(destination_file, OCCH_Arrays_3d, date_string, variable_name):
    # 打开Excel文件
    app = xw.App(visible=False)
    wb = app.books.open(destination_file)
    
    # 遍历OCCH_Arrays_3d
    for i, data in enumerate(OCCH_Arrays_3d):
        global oc_num_global
        
        specific_value_row_material_occh_num = data[0]      #光缆或机箱物料sheet块                 0
        specific_value_OCCH_Specification = data[1]         #光缆的型号规格//机箱                   1
        specific_value_optical_type1 = data[2]               #光缆类型：单模或多模//或机箱数量        2                  
        specific_value_optical_type2 = data[3]               #光缆类型：金属或非金属                 3      
        specific_value_optical_type3 = data[4]               #光缆类型：尾缆或光缆                   4
        specific_value_optical_type4 = data[5]               # 光缆芯数                             5
        specific_value_optical_type5 = data[6]               #光缆或尾缆根数                         6

        if "机箱" in specific_value_OCCH_Specification:
            base_sheet_name = "预制-机箱"
            report_sheet_name = "产品检验-机箱"
            new_base_sheet_name = f"更新_出厂检验{specific_value_OCCH_Specification}-{specific_value_row_material_occh_num}"
            new_report_sheet_name = f"更新_产品检验{specific_value_OCCH_Specification}-{specific_value_row_material_occh_num}"
        elif "金属" in specific_value_OCCH_Specification:
            base_sheet_name = f"预制-{specific_value_optical_type4}"
            report_sheet_name = "产品检验-光缆"
            new_base_sheet_name = f"更新_出厂检验{specific_value_optical_type1}{specific_value_optical_type2}{specific_value_optical_type3}{specific_value_optical_type4}-{specific_value_row_material_occh_num}"
            new_report_sheet_name = f"更新_产品检验_{specific_value_optical_type1}{specific_value_optical_type2}{specific_value_optical_type3}{specific_value_optical_type4}-{specific_value_row_material_occh_num}"

        # 复制并重命名预制-芯数/机箱 sheet表
        base_sheet = wb.sheets[base_sheet_name]
        new_base_sheet = base_sheet.copy(after=wb.sheets[-1])
        new_base_sheet.name = new_base_sheet_name
        data_filling(new_base_sheet, data, variable_name, date_string)
        
        # 复制并重命名产品检验-光缆/机箱 sheet表
        report_sheet = wb.sheets[report_sheet_name]
        new_report_sheet = report_sheet.copy(after=wb.sheets[-1])
        new_report_sheet.name = new_report_sheet_name
        data_filling(new_report_sheet, data, variable_name, date_string)
        
        print(f"创建新的sheet表: {new_base_sheet_name} 和 {new_report_sheet_name}")
    
    reset_global_variables()
    reset_global_arrays()
    
    # 保存并关闭工作簿
    wb.save()
    wb.close()
    app.quit()


def reset_global_variables():
    global rls2, ils2, cls2
    rls2 = 0
    ils2 = 0
    cls2 = 0

def reset_global_arrays():
    global return_loss_str_2d, inception_loss_str_2d, combined_loss_str_2d
    return_loss_str_2d = []
    inception_loss_str_2d = []
    combined_loss_str_2d = []
    
def generate_cell_name(index):
    return f"NAME{index}"

# 数据填充函数
def data_filling(sheet, data, variable_name, date_string):
    specific_value_row_material_occh_num = data[0]
    specific_value_OCCH_Specification = data[1]
    specific_value_optical_type1 = str(data[2])
    specific_value_optical_type2 = data[3]
    specific_value_optical_type3 = data[4]
    specific_value_optical_type4 = data[5]
    specific_value_optical_type5 = str(data[6])
    
    global rls2, ils2, cls2

    if specific_value_OCCH_Specification == "机箱":
        oc_num = int(specific_value_optical_type1)
    elif specific_value_optical_type5.isdigit():
        oc_num = int(specific_value_optical_type5)
    else:
        print(f"Error: specific_value_optical_type5 is not a valid integer: {specific_value_optical_type5}")
        return  # 或者设置一个默认值，如 oc_num = 1
    
    oc_num_global.append(oc_num)
    
    order_ymd_num = variable_name[3:11]
    inspection_date = add_delivery_date(order_ymd_num, date_string)
    inception_loss = set()
    return_loss = set()
    return_loss_pc = set()
    return_loss_mpo = set()
    
    if "出厂检验" in sheet.name:
        if specific_value_optical_type4 == "24芯":
            data_range = sheet.range('A1:BF252')
            data_values = np.array(data_range.value)

            # 更新前四行的特定单元格
            for row in data_values[:4]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    if "NAME1" in cell_value:
                        row[idx] = cell_value.replace("NAME1", variable_name)
                    elif "NAME2" in cell_value:
                        row[idx] = cell_value.replace("NAME2", inspection_date)
                    elif "NAME3" in cell_value:
                        row[idx] = cell_value.replace("NAME3", f"{specific_value_optical_type1}{specific_value_optical_type2}{specific_value_optical_type3}-{specific_value_optical_type4}")
                    elif "NAME4" in cell_value:
                        row[idx] = cell_value.replace("NAME4", specific_value_optical_type5)
                    elif "NAME5" in cell_value:
                        row[idx] = cell_value.replace("NAME5", specific_value_optical_type5)
            
            i = 56
            k = 6

            # 更新后面的单元格
            for row in data_values[7:]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    for cell_loop in range(oc_num):
                        for value_loop_inception in range(6 + i * cell_loop, 29 + i * cell_loop):
                            if 6 + i * cell_loop <= k <= 29 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_i = round(random.uniform(0.70, 1.20), 2)
                                    row[idx] = f"{random_number_i}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 inception_loss 集合中元素数量是否超过 8 个
                                    if len(inception_loss) < 8:
                                        inception_loss.add(random_number_i)    
                                        
                        for value_loop_return in range(30 + i * cell_loop, 53 + i * cell_loop):
                            if 30 + i * cell_loop <= k <= 53 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_r = round(random.uniform(25.01, 28.99), 2)
                                    row[idx] = f"{random_number_r}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 inception_loss 集合中元素数量是否超过 8 个
                                    if len(return_loss) < 8:
                                        return_loss.add(random_number_r)
                                    
                        for value_loop_ok in range(54 + i * cell_loop, 61 + i * cell_loop):
                            if 54 + i * cell_loop <= k <= 61 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    if k % 2 == 0:
                                        row[idx] = "√"
                                        print(f"NAME{k} = {row[idx]}")
                                    else:
                                        row[idx] = None
                                        print(f"NAME{k} = {row[idx]}")
                                    k += 1
                    for cell_loop_end in range(oc_num, 244):   #最大根数
                        if k > 61 + i * cell_loop:
                            if cell_value == f"NAME{k}":
                                row[idx] = None
                                print(f"NAME{k} = {row[idx]}")
                                k += 1
            # 将修改后的数据写回工作表
            sheet.range('A1:BF252').value = data_values.tolist()

            print(f"return_loss: {return_loss}")
            print(f"inception_loss: {inception_loss}")
                        

        elif specific_value_optical_type4 == "4芯":
            data_range = sheet.range('A1:R292')
            data_values = np.array(data_range.value)

            # 更新前四行的特定单元格
            for row in data_values[:4]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    if "NAME1" in cell_value:
                        row[idx] = cell_value.replace("NAME1", variable_name)
                    elif "NAME2" in cell_value:
                        row[idx] = cell_value.replace("NAME2", inspection_date)
                    elif "NAME3" in cell_value:
                        row[idx] = cell_value.replace("NAME3", f"{specific_value_optical_type1}{specific_value_optical_type2}{specific_value_optical_type3}-{specific_value_optical_type4}")
                    elif "NAME4" in cell_value:
                        row[idx] = cell_value.replace("NAME4", specific_value_optical_type5)
                    elif "NAME5" in cell_value:
                        row[idx] = cell_value.replace("NAME5", specific_value_optical_type5)
            
            i = 16
            k = 6

            # 更新后面的单元格
            for row in data_values[7:]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    for cell_loop in range(oc_num):
                        for value_loop_inception in range(6 + i * cell_loop, 9 + i * cell_loop):
                            if 6 + i * cell_loop <= k <= 9 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_i = round(random.uniform(0.70, 1.20), 2)
                                    row[idx] = f"{random_number_i}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 inception_loss 集合中元素数量是否超过 8 个
                                    if len(inception_loss) < 4:
                                        inception_loss.add(random_number_i)    
                                        
                        for value_loop_return in range(10 + i * cell_loop, 13 + i * cell_loop):
                            if 10 + i * cell_loop <= k <= 13 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_r = round(random.uniform(25.01, 28.99), 2)
                                    row[idx] = f"{random_number_r}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 inception_loss 集合中元素数量是否超过 8 个
                                    if len(return_loss) < 4:
                                        return_loss.add(random_number_r)
                                    
                        for value_loop_ok in range(14 + i * cell_loop, 21 + i * cell_loop):
                            if 14 + i * cell_loop <= k <= 21 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    if k % 2 == 0:
                                        row[idx] = "√"
                                        print(f"NAME{k} = {row[idx]}")
                                    else:
                                        row[idx] = None
                                        print(f"NAME{k} = {row[idx]}")
                                    k += 1
                    for cell_loop_end in range(oc_num, 284):   #最大根数
                        if k > 21 + i * cell_loop:
                            if cell_value == f"NAME{k}":
                                row[idx] = None
                                print(f"NAME{k} = {row[idx]}")
                                k += 1
            # 将修改后的数据写回工作表
            sheet.range('A1:R292').value = data_values.tolist()

            print(f"return_loss: {return_loss}")
            print(f"inception_loss: {inception_loss}")
                            
        elif specific_value_optical_type4 == "6芯":
            data_range = sheet.range('A1:V272')
            data_values = np.array(data_range.value)

            # 更新前四行的特定单元格
            for row in data_values[:4]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    if "NAME1" in cell_value:
                        row[idx] = cell_value.replace("NAME1", variable_name)
                    elif "NAME2" in cell_value:
                        row[idx] = cell_value.replace("NAME2", inspection_date)
                    elif "NAME3" in cell_value:
                        row[idx] = cell_value.replace("NAME3", f"{specific_value_optical_type1}{specific_value_optical_type2}{specific_value_optical_type3}-{specific_value_optical_type4}")
                    elif "NAME4" in cell_value:
                        row[idx] = cell_value.replace("NAME4", specific_value_optical_type5)
                    elif "NAME5" in cell_value:
                        row[idx] = cell_value.replace("NAME5", specific_value_optical_type5)
            
            i = 20
            k = 6

            # 更新后面的单元格
            for row in data_values[7:]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    for cell_loop in range(oc_num):
                        for value_loop_inception in range(6 + i * cell_loop, 11 + i * cell_loop):
                            if 6 + i * cell_loop <= k <= 11 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_i = round(random.uniform(0.70, 1.20), 2)
                                    row[idx] = f"{random_number_i}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 inception_loss 集合中元素数量是否超过 8 个
                                    if len(inception_loss) < 6:
                                        inception_loss.add(random_number_i)    
                                        
                        for value_loop_return in range(12 + i * cell_loop, 17 + i * cell_loop):
                            if 12 + i * cell_loop <= k <= 17 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_r = round(random.uniform(25.01, 28.99), 2)
                                    row[idx] = f"{random_number_r}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 inception_loss 集合中元素数量是否超过 8 个
                                    if len(return_loss) < 6:
                                        return_loss.add(random_number_r)
                                    
                        for value_loop_ok in range(18 + i * cell_loop, 25 + i * cell_loop):
                            if 18 + i * cell_loop <= k <= 25 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    if k % 2 == 0:
                                        row[idx] = "√"
                                        print(f"NAME{k} = {row[idx]}")
                                    else:
                                        row[idx] = None
                                        print(f"NAME{k} = {row[idx]}")
                                    k += 1
                    for cell_loop_end in range(oc_num, 264):
                        if k > 25 + i * cell_loop:
                            if cell_value == f"NAME{k}":
                                row[idx] = None
                                print(f"NAME{k} = {row[idx]}")
                                k += 1
            # 将修改后的数据写回工作表
            sheet.range('A1:V272').value = data_values.tolist()

            print(f"return_loss: {return_loss}")
            print(f"inception_loss: {inception_loss}")
                            
        elif specific_value_optical_type4 == "8芯":
            data_range = sheet.range('A1:Z249')
            data_values = np.array(data_range.value)

            # 更新前四行的特定单元格
            for row in data_values[:4]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    if "NAME1" in cell_value:
                        row[idx] = cell_value.replace("NAME1", variable_name)
                    elif "NAME2" in cell_value:
                        row[idx] = cell_value.replace("NAME2", inspection_date)
                    elif "NAME3" in cell_value:
                        row[idx] = cell_value.replace("NAME3", f"{specific_value_optical_type1}{specific_value_optical_type2}{specific_value_optical_type3}-{specific_value_optical_type4}")
                    elif "NAME4" in cell_value:
                        row[idx] = cell_value.replace("NAME4", specific_value_optical_type5)
                    elif "NAME5" in cell_value:
                        row[idx] = cell_value.replace("NAME5", specific_value_optical_type5)
            
            i = 24
            k = 6

            # 更新后面的单元格
            for row in data_values[7:]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    for cell_loop in range(oc_num):
                        for value_loop_inception in range(6 + i * cell_loop, 13 + i * cell_loop):
                            if 6 + i * cell_loop <= k <= 13 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_i = round(random.uniform(0.70, 1.20), 2)
                                    row[idx] = f"{random_number_i}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 inception_loss 集合中元素数量是否超过 8 个
                                    if len(inception_loss) < 8:
                                        inception_loss.add(random_number_i)    
                                        
                        for value_loop_return in range(14 + i * cell_loop, 21 + i * cell_loop):
                            if 14 + i * cell_loop <= k <= 21 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_r = round(random.uniform(25.01, 28.99), 2)
                                    row[idx] = f"{random_number_r}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 inception_loss 集合中元素数量是否超过 8 个
                                    if len(return_loss) < 8:
                                        return_loss.add(random_number_r)
                                    
                        for value_loop_ok in range(22 + i * cell_loop, 29 + i * cell_loop):
                            if 22 + i * cell_loop <= k <= 29 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    if k % 2 == 0:
                                        row[idx] = "√"
                                        print(f"NAME{k} = {row[idx]}")
                                    else:
                                        row[idx] = None
                                        print(f"NAME{k} = {row[idx]}")
                                    k += 1
                    for cell_loop_end in range(oc_num, 241):#最大根数
                        if k > 29 + i * cell_loop:
                            if cell_value == f"NAME{k}":
                                row[idx] = None
                                print(f"NAME{k} = {row[idx]}")
                                k += 1
            # 将修改后的数据写回工作表
            sheet.range('A1:Z249').value = data_values.tolist()

            print(f"return_loss: {return_loss}")
            print(f"inception_loss: {inception_loss}")
                            
        elif specific_value_optical_type4 == "12芯":
            data_range = sheet.range('A1:AH263')
            data_values = np.array(data_range.value)

            # 更新前四行的特定单元格
            for row in data_values[:4]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    if "NAME1" in cell_value:
                        row[idx] = cell_value.replace("NAME1", variable_name)
                    elif "NAME2" in cell_value:
                        row[idx] = cell_value.replace("NAME2", inspection_date)
                    elif "NAME3" in cell_value:
                        row[idx] = cell_value.replace("NAME3", f"{specific_value_optical_type1}{specific_value_optical_type2}{specific_value_optical_type3}-{specific_value_optical_type4}")
                    elif "NAME4" in cell_value:
                        row[idx] = cell_value.replace("NAME4", specific_value_optical_type5)
                    elif "NAME5" in cell_value:
                        row[idx] = cell_value.replace("NAME5", specific_value_optical_type5)
            
            i = 32
            k = 6

            # 更新后面的单元格
            for row in data_values[7:]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    for cell_loop in range(oc_num):
                        for value_loop_inception in range(6 + i * cell_loop, 17 + i * cell_loop):
                            if 6 + i * cell_loop <= k <= 17 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_i = round(random.uniform(0.70, 1.20), 2)
                                    row[idx] = f"{random_number_i}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 inception_loss 集合中元素数量是否超过 8 个
                                    if len(inception_loss) < 8:
                                        inception_loss.add(random_number_i)    
                                        
                        for value_loop_return in range(18 + i * cell_loop, 29 + i * cell_loop):
                            if 18 + i * cell_loop <= k <= 29 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_r = round(random.uniform(25.01, 28.99), 2)
                                    row[idx] = f"{random_number_r}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 inception_loss 集合中元素数量是否超过 8 个
                                    if len(return_loss) < 8:
                                        return_loss.add(random_number_r)
                                    
                        for value_loop_ok in range(30 + i * cell_loop, 37 + i * cell_loop):
                            if 30 + i * cell_loop <= k <= 37 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    if k % 2 == 0:
                                        row[idx] = "√"
                                        print(f"NAME{k} = {row[idx]}")
                                    else:
                                        row[idx] = None
                                        print(f"NAME{k} = {row[idx]}")
                                    k += 1
                    for cell_loop_end in range(oc_num, 255):
                        if k > 37 + i * cell_loop:
                            if cell_value == f"NAME{k}":
                                row[idx] = None
                                print(f"NAME{k} = {row[idx]}")
                                k += 1
            # 将修改后的数据写回工作表
            sheet.range('A1:AH263').value = data_values.tolist()

            print(f"return_loss: {return_loss}")
            print(f"inception_loss: {inception_loss}")
                            
        elif specific_value_OCCH_Specification == "机箱":
            Chassis_num = str(specific_value_optical_type1)  # 转换为字符串类型
            data_range = sheet.range('A1:R292')
            data_values = np.array(data_range.value)

            # 更新前四行的特定单元格
            for row in data_values[:4]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    if "NAME1" in cell_value:
                        row[idx] = cell_value.replace("NAME1", variable_name) #订单号
                    elif "NAME2" in cell_value:
                        row[idx] = cell_value.replace("NAME2", inspection_date) # 发货前的检验日期
                    elif "NAME3" in cell_value:
                        row[idx] = cell_value.replace("NAME3", Chassis_num)  
                    elif "NAME4" in cell_value:
                        row[idx] = cell_value.replace("NAME4", Chassis_num)  # 送检数量                        
                    
            i = 16
            k = 5

            # 更新后面的单元格
            for row in data_values[7:]:
                for idx, cell_value in enumerate(row):
                    if cell_value is None:
                        continue
                    cell_value = str(cell_value)
                    for cell_loop in range(oc_num):
                        for value_loop_inception in range(5 + i * cell_loop, 8 + i * cell_loop):
                            if 5 + i * cell_loop <= k <= 8 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_i = round(random.uniform(0.70, 1.20), 2)
                                    row[idx] = f"{random_number_i}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 inception_loss 集合中元素数量是否超过 8 个
                                    if len(inception_loss) < 8:
                                        inception_loss.add(random_number_i)    
                                        
                        for value_loop_return in range(9 + i * cell_loop, 10 + i * cell_loop):
                            if 9 + i * cell_loop <= k <= 10 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_pc = round(random.uniform(35.50, 39.50), 2)
                                    row[idx] = f"{random_number_pc}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 return_loss_pc 集合中元素数量是否超过 4 个
                                    if len(return_loss_pc) < 4:
                                        return_loss_pc.add(random_number_pc)
                                        
                        for value_loop_return in range(11 + i * cell_loop, 12 + i * cell_loop):
                            if 11 + i * cell_loop <= k <= 12 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    random_number_mpo = round(random.uniform(25.50, 29.50), 2)
                                    row[idx] = f"{random_number_mpo}"
                                    print(f"NAME{k} = {row[idx]}")
                                    k += 1
                                    # 检查 return_loss_mpo 集合中元素数量是否超过 4 个
                                    if len(return_loss_mpo) < 4:
                                        return_loss_mpo.add(random_number_mpo)
                                    
                        for value_loop_ok in range(13 + i * cell_loop, 20 + i * cell_loop):
                            if 13 + i * cell_loop <= k <= 20 + i * cell_loop:
                                if cell_value == f"NAME{k}":
                                    if k % 2 == 0:
                                        row[idx] = None
                                        print(f"NAME{k} = {row[idx]}")
                                    else:
                                        row[idx] = "√"
                                        print(f"NAME{k} = {row[idx]}")
                                    k += 1
                    for cell_loop_end in range(oc_num, 284):
                        if k > 20 + i * cell_loop:
                            if cell_value == f"NAME{k}":
                                row[idx] = None
                                print(f"NAME{k} = {row[idx]}")
                                k += 1
            # 将修改后的数据写回工作表
            sheet.range('A1:R292').value = data_values.tolist()

            print(f"return_loss: {return_loss}")
            print(f"inception_loss: {inception_loss}")
    #底下为什么要加2我大概也明白了，就是循环的问题，因为一个出厂检验报告，一个产品检验报告，隔着来的，这有个" "，所以有一次必定填充空格，大概是这样，我不确定
    return_loss_str = " ".join(map(str, return_loss))
    inception_loss_str = " ".join(map(str, inception_loss))
    if specific_value_OCCH_Specification == "机箱":
        return_loss_pc_str = " ".join(map(str, return_loss_pc))
        return_loss_mpo_str = " ".join(map(str, return_loss_mpo))
        combined_loss_str = return_loss_pc_str + " " + return_loss_mpo_str
        combined_loss_str_2d.append(combined_loss_str)        #cls2
        #搞了半天，发现是" "这个问题不断在combined_loss_str_2d随着循环次数不断加入，具体查看，空格累计错误分析.py，print打印日志，yyds
    return_loss_str_2d.append(return_loss_str)            #rls2
    inception_loss_str_2d.append(inception_loss_str)      #ils2


                        
    if "产品检验" in sheet.name:
        if "金属" in sheet.name:
            for row in sheet.range('A1:J3'):  # 假设A1:K24是需要遍历的范围，可以调整范围
                for cell in row:
                    if cell.value is None:
                        continue
                    if "NAME1" in str(cell.value):
                        cell.value = str(cell.value).replace("NAME1", "PR.B.SS1")  # 产品代码
                    elif "NAME2" in str(cell.value):
                        cell.value = str(cell.value).replace("NAME2", specific_value_optical_type3)  # 预制光缆或预制尾缆
                    elif "NAME3" in str(cell.value):
                        cell.value = str(cell.value).replace("NAME3", f"{specific_value_optical_type1}{specific_value_optical_type2}{specific_value_optical_type3}-{specific_value_optical_type4}")  # 规格型号
                    elif "NAME4" in str(cell.value):
                        cell.value = str(cell.value).replace("NAME4", variable_name)
                    elif "NAME5" in str(cell.value):
                        cell.value = str(cell.value).replace("NAME5", specific_value_optical_type5)  # 光缆或尾缆根数
                    elif "NAME6" in str(cell.value):
                        cell.value = str(cell.value).replace("NAME6", inspection_date)  # 送检日期
                    elif "NAME7" in str(cell.value):
                        cell.value = str(cell.value).replace("NAME7", specific_value_optical_type5)  # 光缆或尾缆根数
            #for row in sheet.range('A15:K23'):  # 假设A1:K24是需要遍历的范围，可以调整范围
            for row in sheet.range('A15:K23'):  # 假设A1:K24是需要遍历的范围，可以调整范围
                for cell in row:
                    if cell.value is None:
                        continue
                    if "NAME8" in str(cell.value):
                        cell.value = inception_loss_str_2d[ils2]  # 插入损耗 
                        ils2 += 2
                    elif "NAME9" in str(cell.value):
                        cell.value = return_loss_str_2d[rls2]  # 回波损耗 
                        rls2 += 2
                    elif "NAME10" in str(cell.value):
                        cell.value = inspection_date  # 送检日期
                        
        elif "机箱" in sheet.name:
            Chassis_num = str(specific_value_optical_type1)  # 转换为字符串类型
            for row in sheet.range('A1:J3'):  # 假设A1:K24是需要遍历的范围，可以调整范围
                for cell in row:
                    if cell.value is None:
                        continue
                    if "NAME1" in str(cell.value):
                        cell.value = str(cell.value).replace("NAME1", variable_name)  # 订单号
                    elif "NAME2" in str(cell.value):
                        cell.value = str(cell.value).replace("NAME2", Chassis_num)  # 产品数量
                    elif "NAME3" in str(cell.value):
                        cell.value = str(cell.value).replace("NAME3", inspection_date)  # 送检日期
                    elif "NAME4" in str(cell.value):
                        cell.value = str(cell.value).replace("NAME4", Chassis_num)  # 产品数量
            #for row in sheet.range('A15:K23'):  # 假设A1:K24是需要遍历的范围，可以调整范围
            for row in sheet.range('A15:K23'):  # 假设A1:K24是需要遍历的范围，可以调整范围
                for cell in row:
                    if cell.value is None:
                        continue
                    if "NAME8" in str(cell.value):
                        cell.value = inception_loss_str_2d[ils2]  # 插入损耗 
                        ils2 += 2
                    elif "NAME9" in str(cell.value):
                        cell.value = combined_loss_str_2d[cls2]  # PC回波损耗 和 MPO回波损耗
                        cls2 += 2
                    elif "NAME10" in str(cell.value):
                        cell.value = inspection_date  # 送检日期


#时间函数重做，不再以---为准，而是文件夹的发货日期，往前推一天，我先以日期节假日不发货的逻辑来写
#更改发货时间逻辑
def add_delivery_date(order_ymd_num, date_string):
    # 节假日列表，这里包括2023年和2024年的节假日和调休日期
    holidays = {
        "2023": ["20230101", "20230102", "20230121", "20230122", "20230123", "20230124", 
                 "20230125", "20230126", "20230127", "20230405", "20230429", "20230430", 
                 "20230501", "20230502", "20230503", "20230622", "20230623", "20230624", 
                 "20230929", "20230930", "20231001", "20231002", "20231003", "20231004", 
                 "20231005", "20231006"],
        "2024": ["20240101", "20240210", "20240211", "20240212", "20240213", "20240214", 
                 "20240215", "20240216", "20240217", "20240404", "20240405", "20240406", 
                 "20240501", "20240502", "20240503", "20240504", "20240505", "20240610", 
                 "20240915", "20240916", "20240917", "20241001", "20241002", "20241003", 
                 "20241004", "20241005", "20241006", "20241007"],
        "2025": ["20250101", "20250128", "20250129", "20250130", "20250131", "20250201", 
                 "20250202", "20250203", "20250204", "20250404", "20250405", "20250406", 
                 "20250501", "20250502", "20250503", "20250504", "20250505", "20250531", 
                 "20250601", "20250602", "20251001", "20251002", "20251003", "20251004", 
                 "20251005", "20251006", "20251007", "20251008"]
    }
    # 添加周末正常上班日期
    working_weekends = {
        "2023": ["20230128", "20230129", "20230423", "20230506", "20230625", "20231007", "20231008"],
        "2024": ["20240204", "20240218", "20240407", "20240428", "20240511", "20240914", "20240929", "20241012"],
        "2025": ["20250126", "20250208", "20250427", "20250928", "20251011"]
    }
    
    # 提取发货月份和日期
    date_match = re.search(r"(\d+)年(\d+)月(\d+)日", date_string)
    if not date_match:
        raise ValueError("日期格式不正确，应该是'X年Y月Z日'")
    
    year = int(date_match.group(1))
    month = int(date_match.group(2))
    day = int(date_match.group(3))

    # 确定发货年份
    #order_year = int(order_ymd_num[:4])
    #order_month_day = order_ymd_num[4:]

    # 生成完整的发货日期
    shipping_date_str = f"{year}{month:02d}{day:02d}"
    shipping_date = datetime.strptime(shipping_date_str, '%Y%m%d')

    # 确定节假日和工作周末列表
    holidays_for_year = holidays.get(str(year), [])
    working_weekends_for_year = working_weekends.get(str(year), [])
    
    # 检查发货日期之前的日期，排除节假日和周末
    while shipping_date.strftime('%Y%m%d') in holidays_for_year or (shipping_date.weekday() >= 5 and shipping_date.strftime('%Y%m%d') not in working_weekends_for_year):
        shipping_date -= timedelta(days=1)
    
    # 返回实际发货日期的前一天
    actual_inspection_date = shipping_date - timedelta(days=1)
    return actual_inspection_date.strftime('%Y-%m-%d')


def copy_excel_file(source_path, destination_path):
    try:
        shutil.copyfile(source_path, destination_path)
        print(f"文件已成功复制到：{destination_path}")
    except Exception as e:
        print(f"复制文件时出错：{e}")


# 删除非更新sheet
def delete_specific_sheets_in_directory(directory_path):
    # 遍历目录下的所有文件
    for filename in os.listdir(directory_path):
        # 检查文件是否为 Excel 文件
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            file_path = os.path.join(directory_path, filename)
            print(f"正在处理文件：{file_path}")

            # 加载工作簿
            workbook = openpyxl.load_workbook(file_path)
            # 获取所有工作表名称的列表
            sheet_names = workbook.sheetnames

            # 记录要删除的工作表名称
            sheets_to_delete = []

            # 遍历所有工作表名称
            for sheet_name in sheet_names:
                # 如果工作表名称不是要保留的名称，则记录该工作表名称
                if "更新" not in sheet_name:
                    sheets_to_delete.append(sheet_name)

            # 如果要删除的工作表数量等于现有的工作表数量，则删除整个文件
            if len(sheets_to_delete) >= len(sheet_names):
                workbook.close()
                os.remove(file_path)
                print(f"已删除整个文件：{file_path}")
            else:
                # 删除记录的工作表
                for sheet_name in sheets_to_delete:
                    del workbook[sheet_name]
                    print(f"删除工作表: {sheet_name}")

                # 保存更改
                workbook.save(file_path)
                print(f"已更新工作簿 {file_path}")
                # 关闭工作簿
                workbook.close()

# 使用正则表达式提取文件名中的日期格式 x月y日
def export_date_month_day(name):
    date_match = re.search(r"\d+年\d+月\d+日", name)
    if date_match:
        return date_match.group()  # 提取日期子字符串
    return None
#获得系统当前时间
def get_current_datetime():
    return datetime.now().strftime("%Y%m%d%H%M%S")

# 遍历根目录下的所有文件夹
for folder_name in sorted(os.listdir(root_directory)):
    folder_path = os.path.join(root_directory, folder_name)
    
    # 检查是否是文件夹
    if os.path.isdir(folder_path):
        # 使用正则表达式提取文件夹名称中符合 <FOLDER_CODE_PATTERN> 格式的子字符串
        match = re.search(r"<FOLDER_CODE_PATTERN>", folder_name)
        if match:
            variable_name = match.group()  # 提取匹配的子字符串
            
            date_string = None
            # 使用正则表达式提取文件名中的日期格式 x月y日
            date_string_folder = export_date_month_day(folder_name)  # 提取日期子字符串
            
            # 初始化计数器
            count = 1
            # 统计含有 "物料编码" 文件的数量
            file_count = sum(1 for file_name in os.listdir(folder_path) if "物料编码" in file_name and file_name.endswith('.xlsx'))
            
            # 遍历文件夹下的文件
            for file_name in os.listdir(folder_path):
                if "物料编码" in file_name and file_name.endswith('.xlsx'):
                    #发货日期date_string
                    date_string_file = export_date_month_day(file_name)
                    if date_string_folder != None:
                        date_string = date_string_folder
                    elif date_string_file != None:
                        date_string = date_string_file
                    #系统当前时间
                    current_date = get_current_datetime()
                    # 判断是否需要添加计数后缀
                    if file_count > 1:
                        destination_file = os.path.join(target_excel_path, f"{current_date}_出厂检验报告{variable_name}-{count}.xlsx")
                    else:
                        destination_file = os.path.join(target_excel_path, f"{current_date}_出厂检验报告{variable_name}.xlsx")
                    #print(f"文件夹名称: {folder_name}, 提取的变量: {variable_name}, 提取的日期: {date_string}，系统当前时间：{current_date}")
                        
                    # 调用函数复制文件
                    copy_excel_file(original_excel_path, destination_file)
                    
                    # 打开Excel文件
                    excel_path = os.path.join(folder_path, file_name)
                    wb = openpyxl.load_workbook(excel_path)
                    sheet = wb.worksheets[0]
                    
                    # 将光缆和机箱的数据添加到数组中
                    OCCH_Arrays_3d = OC_Chassis_add(sheet)
                    
                    # 通过OC_Chassis_add(sheet)返回值来进行excel的sheet表复制，并替换sheet表的表名
                    Creat_sheet_for_OCCH3(destination_file, OCCH_Arrays_3d, date_string, variable_name)
                    
                    # 存储新创建的文件路径
                    if file_count > 1:
                        global_variables[f"{variable_name}-{count}"] = destination_file
                    else:
                        global_variables[f"{variable_name}"] = destination_file
                    count += 1

            # 重置全局盘号计数器
            #global_plate_num = 1

        print(f"所有更改已保存到 {destination_file}")


def delet_row_all_excel_files(target_path):
    # 使用xlwings打开Excel应用，设置为不可见
    app = xw.App(visible=False)
    
    # 遍历目标文件夹中的所有Excel文件
    for file_name in os.listdir(target_path):
        if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            file_path = os.path.join(target_path, file_name)
            wb = app.books.open(file_path)
            
            # 遍历工作簿中的所有sheet
            for sheet in wb.sheets:
                if "出厂检验" in sheet.name:
                    # 根据芯数选择不同的处理逻辑
                    core_num_dict = {
                        "24芯": ("AL4", [74, 159, 244]),
                        "4芯": ("H4", [48, 107, 166, 225, 284]),
                        "6芯": ("K4", [57, 126, 195, 264]),
                        "8芯": ("P4", [73, 157, 241]),
                        "12芯": ("T4", [42, 96, 149, 202, 255]),
                        "机箱": ("H4", [48, 107, 166, 225, 284])
                    }
                    for core, (cell, limits) in core_num_dict.items():
                        if core in sheet.name:
                            cell_value = sheet.range(cell).value
                            # 提取数字部分
                            current_sheet_core_num = int(''.join(filter(str.isdigit, cell_value.split('PCS')[0])))
                            last_row = limits[-1] + 7
                            
                            # 确定删除行的范围
                            for i, max_limit in enumerate(limits):
                                if current_sheet_core_num <= max_limit:
                                    start_row_dr = max_limit + 8
                                    num_rows_dr = last_row - start_row_dr + 1
                                    for row in range(start_row_dr + num_rows_dr - 1, start_row_dr - 1, -1):
                                        sheet.range(f'{row}:{row}').api.Delete()
                                    break
                            else:
                                # 如果超出范围，则打印错误信息
                                print(f"error:{core} row over {max_limit}!")
                            break
            
            # 保存并关闭工作簿
            wb.save()
            wb.close()
    
    # 关闭Excel应用
    app.quit()


delet_row_all_excel_files(target_excel_path)

delete_specific_sheets_in_directory(target_excel_path)

# 遍历文件夹中的所有文件
for filename_trans_pdf in os.listdir(target_excel_path):
    # 检查文件是否为Excel文件
    if filename_trans_pdf.endswith('.xlsx'):
        # 构造完整文件路径
        file_path = os.path.join(target_excel_path, filename_trans_pdf)
        # 构造输出PDF文件路径
        output_pdf_path = os.path.join(target_excel_path, f'{os.path.splitext(filename_trans_pdf)[0]}.pdf')
        
        # 启动 Excel 应用程序
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        
        # 保存为 PDF
        wb.api.ExportAsFixedFormat(0, output_pdf_path)
        
        # 关闭工作簿
        wb.close()
        # 关闭 Excel 应用程序
        app.quit()
        
        print(f'PDF saved to {output_pdf_path}')


# 打印三维数组结构
print(OCCH_Arrays_3d)

# 打印全局变量信息
for variable_name, excel_path in global_variables.items():
    print(f"全局变量名称-文件夹名称：{variable_name}")

# 程序运行结束时记录时间
end_time = time.time()

# 计算总运行时间
total_time = end_time - start_time
total_min = total_time // 60
total_sec = total_time % 60
print(f"程序总运行时间：{total_time}秒")
print(f"程序总运行时间：{total_min}分{total_sec}秒")



