
import re
import os
import openpyxl
import shutil
import time
import xlwings as xw
from datetime import datetime, timedelta
#  先
# 程序开始时记录当前时间
start_time = time.time()

# 全局变量字典，用于存储文件夹名称和对应的Excel文件
global_variables = {}

# 根目录文件夹路径 
root_directory = r'yourpath' 

# 目标文件夹路径 
target_excel_path = r'yourpath' 

# 原始模板路径 
original_excel_path = r'yourpath'

#货物计算器
def num_calculator(sheet):
    # 从A2开始遍历列A并统计有序的数字行数
    start_row = 2  # 开始行号
    current_value = 1  # 从1开始的数字
    row_count = 0
    goods_num = 0  # 初始化 goods_num 为0

    for row in sheet.iter_rows(min_row=start_row, max_col=1):
        cell = row[0]
        if cell.value == current_value:
            row_count += 1
            row_number = start_row + row_count - 1
            current_value += 1
            cell_value_b = sheet.cell(row=row_number, column=2).value
            
            # 检查G列或H列是否有纯数字
            cell_value_g = sheet.cell(row=row_number, column=7).value
            cell_value_h = sheet.cell(row=row_number, column=8).value
            
            if cell_value_g is not None and str(cell_value_g).isdigit():
                cell_value = int(cell_value_g)
            elif cell_value_h is not None and str(cell_value_h).isdigit():
                cell_value = int(cell_value_h)
            else:
                print(f"第{row_number}行的G列和H列都没有纯数字: G列值: {cell_value_g}, H列值: {cell_value_h}")
                continue  # 如果G列和H列都没有纯数字，跳过该行

            if "光缆" in cell_value_b:
                goods_num += cell_value  # 如果是光缆，累加数量到 goods_num

    return goods_num


#时间模块 节假日则一直推到工作日
def add_delivery_date(start_date_str, goods_num):
    # 节假日列表，这里包括2023年和2024年的节假日和调休日期
    holidays_2023 = ["20230101", "20230102", "20230121", "20230122", "20230123", "20230124", 
                     "20230125", "20230126", "20230127", "20230405", "20230429", "20230430", 
                     "20230501", "20230502", "20230503", "20230622", "20230623", "20230624", 
                     "20230929", "20230930", "20231001", "20231002", "20231003", "20231004", 
                     "20231005", "20231006"]
    holidays_2024 = ["20240101", "20240210", "20240211", "20240212", "20240213", "20240214", 
                     "20240215", "20240216", "20240217", "20240404", "20240405", "20240406", 
                     "20240501", "20240502", "20240503", "20240504", "20240505", "20240610", 
                     "20240915", "20240916", "20240917", "20241001", "20241002", "20241003", 
                     "20241004", "20241005", "20241006", "20241007"]
    holidays_2025 = ["20250101", "20250128", "20250129", "20250130", "20250131", "20250201", 
                     "20250202", "20250203", "20250204", "20250404", "20250405", "20250406", 
                     "20250501", "20250502", "20250503", "20250504", "20250505", "20250531", 
                     "20250601", "20250602", "20251001", "20251002", "20251003", "20251004", 
                     "20251005", "20251006", "20251007", "20251008"]
    # 添加周末正常上班日期
    working_weekends_2023 = ["20230128", "20230129", "20230423", "20230506", "20230625", "20231007", "20231008"]
    working_weekends_2024 = ["20240204", "20240218", "20240407", "20240428", "20240511", "20240914", "20240929", "20241012"]
    working_weekends_2025 = ["20250126", "20250208", "20250427", "20250928", "20251011"]
    # 将输入的日期字符串转换为日期对象
    start_date = datetime.strptime(start_date_str, '%Y%m%d')
    # 判断是2023年还是2024年，并设置相应的节假日和工作周末列表
    year = start_date.year
    if year == 2023:
        holidays = holidays_2023
        working_weekends = working_weekends_2023
    elif year == 2024:
        holidays = holidays_2024
        working_weekends = working_weekends_2024
    elif year == 2025:
        holidays = holidays_2025
        working_weekends = working_weekends_2025
    # 定义一个变量来追踪已经处理的货物数量
    processed_goods = 0
    # 如果是大量订单，增加两天开始处理
    if goods_num >= 10000:
        start_date += timedelta(days=2)
    else:  # 如果是小量订单，增加一天开始处理
        start_date += timedelta(days=1)
    while processed_goods < goods_num:
        # 检查是否为节假日或周末（除非是工作周末）
        if start_date.strftime('%Y%m%d') in holidays or (start_date.weekday() == 6 and start_date.strftime('%Y%m%d') not in working_weekends):
            start_date += timedelta(days=1)
            continue
        break
    # 处理完当前批货物后，增加一天
    start_date += timedelta(days=1)
    # 返回处理完所有货物的日期（前一天）
    return (start_date - timedelta(days=1)).strftime('%Y.%m.%d').replace('.0', '.')

def get_sample_info(x):
    info = ""
    sample_size, ac, re = 0, [], []

    if 1 <= x <= 8:
        info = "批量范围:1~~~8\n样本大小 = 2\nAc: 0  0  0\nRe: 1  1  1\n"
        sample_size, ac, re = 2, [0, 0, 0], [1, 1, 1]
    elif 9 <= x <= 15:
        info = "批量范围:9~~~15\n样本大小 = 3\nAc: 0  0  0\nRe: 1  1  1\n"
        sample_size, ac, re = 3, [0, 0, 0], [1, 1, 1]
    elif 16 <= x <= 25:
        info = "批量范围:16~~~25\n样本大小 = 5\nAc: 0  0  0\nRe: 1  1  1\n"
        sample_size, ac, re = 5, [0, 0, 0], [1, 1, 1]
    elif 26 <= x <= 50:
        info = "批量范围:26~~~50\n样本大小 = 8\nAc: 0  0  0\nRe: 1  1  1\n"
        sample_size, ac, re = 8, [0, 0, 0], [1, 1, 1]
    elif 51 <= x <= 90:
        info = "批量范围:51~~~90\n样本大小 = 13\nAc: 0  0  0\nRe: 1  1  1\n"
        sample_size, ac, re = 13, [0, 0, 0], [1, 1, 1]
    elif 91 <= x <= 150:
        info = "批量范围:91~~~150\n样本大小 = 20\nAc: 0  0  1\nRe: 1  1  2\n"
        sample_size, ac, re = 20, [0, 0, 1], [1, 1, 2]
    elif 151 <= x <= 280:
        info = "批量范围:151~~~280\n样本大小 = 32\nAc: 0  1  1\nRe: 1  2  2\n"
        sample_size, ac, re = 32, [0, 1, 1], [1, 2, 2]
    elif 281 <= x <= 500:
        info = "批量范围:281~~~500\n样本大小 = 50\nAc: 0  1  1\nRe: 1  2  2\n"
        sample_size, ac, re = 50, [0, 1, 1], [1, 2, 2]
    elif 501 <= x <= 1200:
        info = "批量范围:501~~~1200\n样本大小 = 80\nAc: 1  1  2\nRe: 2  2  3\n"
        sample_size, ac, re = 80, [1, 1, 2], [2, 2, 3]
    elif 1201 <= x <= 3200:
        info = "批量范围:1201~~~3200\n样本大小 = 125\nAc: 1  2  3\nRe: 2  3  4\n"
        sample_size, ac, re = 125, [1, 2, 3], [2, 3, 4]
    elif 3201 <= x <= 10000:
        info = "批量范围:3201~~~10000\n样本大小 = 200\nAc: 1  3  5\nRe: 2  4  6\n"
        sample_size, ac, re = 200, [1, 3, 5], [2, 4, 6]
    elif 10001 <= x <= 35000:
        info = "批量范围:10001~~~35000\n样本大小 = 315\nAc: 2  5  7\nRe: 3  6  8\n"
        sample_size, ac, re = 315, [2, 5, 7], [3, 6, 8]
    elif 35001 <= x <= 150000:
        info = "批量范围:35001~~~150000\n样本大小 = 500\nAc: 3  7  10\nRe: 4  8  11\n"
        sample_size, ac, re = 500, [3, 7, 10], [4, 8, 11]
    elif 150001 <= x <= 500000:
        info = "批量范围:150001~~~500000\n样本大小 = 800\nAc: 5  10  14\nRe: 6  11  15\n"
        sample_size, ac, re = 800, [5, 10, 14], [6, 11, 15]
    elif x >= 500001:
        info = "批量范围:≥500001\n样本大小 = 1250\nAc: 7  14  21\nRe: 8  15  22\n"
        sample_size, ac, re = 1250, [7, 14, 21], [8, 15, 22]
    else:
        info = "无效的输入，请输入一个正整数。"
    print(info)
    return sample_size, ac, re

def rename_and_replace_sheet(goods_num,sheet_num, destination_file, old_prefix, new_prefix, cell_value_f, cell_value_c, variable_name, cell_value):
    
    # 在函数开始时打开工作簿
    creat_file_open = openpyxl.load_workbook(destination_file)
    # 构造旧工作表名称
    old_sheet_name = f"{old_prefix}{sheet_num}"
    new_sheet_name = f"{new_prefix}{sheet_num}"

    print(f"正在处理工作表：{old_sheet_name}")
    
    start_date_str = variable_name[3:11]
    extention_time = add_delivery_date(start_date_str, goods_num)
    
    sample_size, ac, re = get_sample_info(cell_value)
    first_ac = ac[0]
    second_ac = ac[1]
    third_ac = ac[2]
    first_re = re[0]
    second_re = re[1]
    third_re = re[2]
    # 检查工作表是否存在
    if old_sheet_name in creat_file_open.sheetnames:
        sheet = creat_file_open[old_sheet_name]

        # 遍历工作表的所有单元格并替换文本
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == None:
                    continue
                #print(f"当前单元格值：{cell.value}")  # 打印当前单元格值
                if cell.value == "NAME1":
                    cell.value = cell_value_f  #品牌
                elif cell.value == "NAME2":
                    cell.value = cell_value_c  #规格型号
                elif cell.value == "NAME3":
                    cell.value = variable_name #需求单号
                elif cell.value == "NAME4":
                    cell.value = cell_value    #产品数量                                   
                elif cell.value == "NAME5":
                    cell.value = extention_time      #送检日期          
                elif cell.value == "NAME6":
                    cell.value = sample_size   #检验数量            
                elif cell.value == "NAME7":    #AC数
                    cell.value = first_ac
                elif cell.value == "NAME8":
                    cell.value = second_ac    
                elif cell.value == "NAME9":
                    cell.value = third_ac
                elif cell.value == "NAME10":
                    cell.value = first_re      #RE数
                elif cell.value == "NAME11":
                    cell.value = second_re
                elif cell.value == "NAME12":
                    cell.value = third_re  
                elif cell.value == "NAME13":
                    cell.value = extention_time       #日期
                
        # 重命名工作表
        sheet.title = new_sheet_name
        print(f"重命名工作表: {old_sheet_name} 为 {new_sheet_name}")
        # 保存更改
        creat_file_open.save(destination_file)
        creat_file_open.close()
        print("更改已保存。")
    else:
        print(f"工作表 {old_sheet_name} 不存在。")
        

def copy_excel_file(source_path, destination_path):
    try:
        shutil.copyfile(source_path, destination_path)
        print(f"文件已成功复制到：{destination_path}")
    except Exception as e:
        print(f"复制文件时出错：{e}")


def number_line(sheet, folder_name, file_name,destination_file,goods_num):
    # 从A8开始遍历列A并统计有序的数字行数
    start_row = 2  # 开始行号
    column_name = 'A'  # 列名
    current_value = 1  # 从1开始的数字
    row_count = 0
    sheet_num_m = 1   #多模sheet表格默认开始序号
    sheet_num_s = 1   #单模sheet表格默认开始序号
    sheet_num_am = 1   #航空插头公头sheet表格默认开始序号
    sheet_num_af = 1   #航空插头母头sheet表格默认开始序号
    sheet_num_t = 1   #机箱sheet表格默认开始序号
    sheet_num_c = 1   #连接器sheet表格默认开始序号
    sheet_num_ado = 1  #适配器单联sheet表格默认开始序号
    sheet_num_add = 1  #适配器双联sheet表格默认开始序号
    sheet_num_adf = 1  #适配器四联sheet表格默认开始序号
    sheet_num_b = 1   #分支器sheet表格默认开始序号
    print(f"文件夹名称：{folder_name}, 文件名：{file_name}, Sheet1（合同）")
        
    for row in sheet.iter_rows(min_row=start_row, max_col=1):
        cell = row[0]
        if cell.value == current_value:  # 如果单元格中的值等于当前值
            row_count += 1    
            row_number = start_row + row_count - 1  # 计算当前行号
            current_value += 1  # 增加当前值以匹配下一个数字
            # 根据不同的值执行不同的操作
            cell_value_b = sheet.cell(row=row_number, column=2).value  # B列的值 名称
            cell_value_c = sheet.cell(row=row_number, column=3).value  # C列的值 规格型号
            
            # 判断G列或H列是否有纯数字
            cell_value_g = sheet.cell(row=row_number, column=7).value
            cell_value_h = sheet.cell(row=row_number, column=8).value
            
            if cell_value_g is not None and str(cell_value_g).isdigit():
                cell_value = int(cell_value_g)
                cell_value_f = sheet.cell(row=row_number, column=5).value  # E列的值 品牌
            elif cell_value_h is not None and str(cell_value_h).isdigit():
                cell_value = int(cell_value_h)
                cell_value_f = sheet.cell(row=row_number, column=6).value  # F列的值 品牌
            else:
                print(f"G or H no value! It's {row_number} row")
                continue  # 如果G列和H列都没有纯数字，跳过该行

            # 判断B列的元素属于哪种产品类型
            if "光缆" in cell_value_b:
                if "多模" in cell_value_c:
                    # 使用函数
                    old_prefix = "多模"
                    new_prefix = "多模更新"
                    sheet_num = sheet_num_m
                    rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                    print(f"{row_count}:多模 H{row_number}数据：{cell_value} 多模sheet{sheet_num_m}")
                    sheet_num_m += 1
                    continue
                elif "单模" in cell_value_c:
                    old_prefix = "单模"
                    new_prefix = "单模更新"
                    sheet_num = sheet_num_s
                    rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                    print(f"{row_count}:单模 H{row_number}数据：{cell_value} 单模sheet{sheet_num_s}")
                    sheet_num_s += 1
                    continue
            elif "航空插头" in cell_value_b:
                if "公头" in cell_value_c:
                    old_prefix = "航空插头公头"
                    new_prefix = "公头更新"
                    sheet_num = sheet_num_am
                    rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                    print(f"{row_count}:航空插头公头 H{row_number}数据：{cell_value} 公头sheet{sheet_num_am}")
                    sheet_num_am += 1
                    continue
                elif "母头" in cell_value_c:
                    old_prefix = "航空插头母头"
                    new_prefix = "母头更新"
                    sheet_num = sheet_num_af
                    rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                    print(f"{row_count}:航空插头母头 H{row_number}数据：{cell_value} 母头sheet{sheet_num_af}")
                    sheet_num_af += 1
                    continue            
            elif "机箱" in cell_value_b:
                if "1U" in cell_value_c:
                    old_prefix = "机箱"
                    new_prefix = "1U机箱更新"
                    sheet_num = sheet_num_t
                    rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                    print(f"{row_count}:机箱1U H{row_number}数据：{cell_value} 机箱1Usheet{sheet_num_t}")
                    sheet_num_t += 1
                    continue
                else:
                    print("有其他类型的机箱")
            elif "连接器" in cell_value_b:
                old_prefix = "连接器"
                new_prefix = "连接器更新"
                sheet_num = sheet_num_c
                rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                print(f"{row_count}:连接器 H{row_number}数据：{cell_value} 连接器sheet{sheet_num_c}")
                sheet_num_c += 1
                continue
            elif "适配器" in cell_value_b:
                if "单联" in cell_value_c:
                    old_prefix = "适配器单联"
                    new_prefix = "适配器单联更新"
                    sheet_num = sheet_num_ado
                    rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                    print(f"{row_count}:适配器单联 H{row_number}数据：{cell_value} 适配器单联sheet{sheet_num_ado}")
                    sheet_num_ado += 1
                    continue
                elif "双联" in cell_value_c:
                    old_prefix = "适配器双联"
                    new_prefix = "适配器双联更新"
                    sheet_num = sheet_num_add
                    rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                    print(f"{row_count}:适配器双联 H{row_number}数据：{cell_value} 适配器双联sheet{sheet_num_add}")
                    sheet_num_add += 1
                    continue     
                elif "四联" in cell_value_c:
                    old_prefix = "适配器四联"
                    new_prefix = "适配器四联更新"
                    sheet_num = sheet_num_adf
                    rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                    print(f"{row_count}:适配器四联 H{row_number}数据：{cell_value} 适配器四联sheet{sheet_num_adf}")
                    sheet_num_adf += 1
                    continue     
            elif "分支器" in cell_value_b:
                old_prefix = "分支器"
                new_prefix = "分支器更新"
                sheet_num = sheet_num_b
                rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                print(f"{row_count}:分支器 H{row_number}数据：{cell_value} 分支器sheet{sheet_num_b}")
                sheet_num_b += 1
                continue
            elif "分支拉拽" in cell_value_b:
                old_prefix = "分支拉拽"
                new_prefix = "分支拉拽更新"
                sheet_num = sheet_num_b
                rename_and_replace_sheet(goods_num,sheet_num,destination_file, old_prefix, new_prefix, cell_value_f,cell_value_c,variable_name,cell_value)
                print(f"{row_count}:分支拉拽 H{row_number}数据：{cell_value} 分支拉拽sheet{sheet_num_b}")
                sheet_num_b += 1
                continue
            else:
                print(f"error line {row_count}") # 分辨是第几行定位错误地址
        else:
            break  # 如果不是有序的数字，停止遍历
    # 打印有序数字的行数
    print(f"从{column_name}{start_row}开始共有{row_count}行有序数字。\n")


#删除非更新sheet
def delete_specific_sheets_in_directory(directory_path):
    # 遍历目录下的所有文件
    for filename in os.listdir(directory_path):
        # 检查文件是否为 Excel 文件
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            file_path = os.path.join(directory_path, filename)
            print(f"正在处理文件：{file_path}")

            # 加载工作簿
            workbook = openpyxl.load_workbook(file_path)
            # 获取所有工作表名称的列表
            sheet_names = workbook.sheetnames
            # 遍历所有工作表名称
            for sheet_name in sheet_names:
                # 如果工作表名称不是要保留的名称，则删除该工作表
                if "更新" not in sheet_name:
                    del workbook[sheet_name]
                    print(f"删除工作表: {sheet_name}")

            # 保存更改
            workbook.save(file_path)
            print(f"已更新工作簿 {file_path}")
            # 关闭工作簿
            workbook.close()

#获得系统当前时间
def get_current_datetime():
    return datetime.now().strftime("%Y%m%d%H%M%S")

# 遍历根目录下的所有文件夹
for folder_name in sorted(os.listdir(root_directory)):
    folder_path = os.path.join(root_directory, folder_name)
    
    # 检查是否是文件夹
    if os.path.isdir(folder_path):
        # 使用正则表达式提取文件夹名称中符合 <CODE_PATTERN> 格式的子字符串
        match = re.search(r"<CODE_PATTERN>", folder_name)
        if match:
            variable_name = match.group()  # 提取匹配的子字符串
            
            # 初始化计数器
            count = 1
            # 统计含有 "物料编码" 文件的数量
            file_count = sum(1 for file_name in os.listdir(folder_path) if "物料编码" in file_name and file_name.endswith('.xlsx'))
            
            # 遍历文件夹下的文件
            for file_name in os.listdir(folder_path):
                if "物料编码" in file_name and file_name.endswith('.xlsx'):
                    #系统当前时间
                    current_date = get_current_datetime()
                    # 判断是否需要添加计数后缀
                    if file_count > 1:
                        destination_file = os.path.join(target_excel_path, f"{current_date}_来料检验{variable_name}-{count}.xlsx")
                    else:
                        destination_file = os.path.join(target_excel_path, f"{current_date}_来料检验{variable_name}.xlsx")
                    
                    # 调用函数复制文件
                    copy_excel_file(original_excel_path, destination_file)
                    # 打开Excel文件
                    excel_path = os.path.join(folder_path, file_name)
                    wb = openpyxl.load_workbook(excel_path)
                    
                    sheet = wb.worksheets[0]
                    # 数量计算，天数要用 <DELAY_QUANTITY> <DELAY_DAYS>
                    goods_num = num_calculator(sheet)
                    # 打印Sheet1（合同）的H8开始直到结束的数据
                    number_line(sheet, folder_name, file_name, destination_file, goods_num)
                    
                    # 在函数结束时保存并关闭工作簿
                    # creat_file_open.save(destination_file)
                    # creat_file_open.close()
                    # 存储新创建的文件路径
                    if file_count > 1:
                        global_variables[f"{variable_name}-{count}"] = destination_file
                    else:
                        global_variables[f"{variable_name}"] = destination_file
                    count += 1
                    
        print(f"所有更改已保存到 {destination_file}")


delete_specific_sheets_in_directory(target_excel_path)

# 遍历文件夹中的所有文件
for filename_trans_pdf in os.listdir(target_excel_path):
    # 检查文件是否为Excel文件
    if filename_trans_pdf.endswith('.xlsx'):
        # 构造完整文件路径
        file_path = os.path.join(target_excel_path, filename_trans_pdf)
        # 构造输出PDF文件路径
        output_pdf_path = os.path.join(target_excel_path, f'{os.path.splitext(filename_trans_pdf)[0]}.pdf')
        
        # 启动 Excel 应用程序
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        
        # 保存为 PDF
        wb.api.ExportAsFixedFormat(0, output_pdf_path)
        
        # 关闭工作簿
        wb.close()
        # 关闭 Excel 应用程序
        app.quit()
        
        print(f'PDF saved to {output_pdf_path}')
        
# 打印全局变量信息
for variable_name, excel_path in global_variables.items():
    print(f"全局变量名称-文件夹名称：{variable_name}")

# 程序运行结束时记录时间
end_time = time.time()

# 计算总运行时间
total_time = end_time - start_time
total_min = total_time // 60
total_sec = total_time % 60
print(f"程序总运行时间：{total_time}秒")
print(f"程序总运行时间：{total_min}分{total_sec}秒")



